import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import requests
from bs4 import BeautifulSoup
import threading
from urllib.parse import urljoin, urlparse
import pandas as pd
from datetime import datetime, timedelta
import os
import time
import re
import unicodedata
import json
import pickle
import smtplib
import email.mime.text
import email.mime.multipart
MimeText = email.mime.text.MIMEText
MimeMultipart = email.mime.multipart.MIMEMultipart
import schedule
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service

# 선택적 import (없어도 프로그램 실행 가능)
try:
    import matplotlib.pyplot as plt
    import matplotlib.font_manager as fm
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    import seaborn as sns
    CHART_AVAILABLE = True
except ImportError:
    CHART_AVAILABLE = False

try:
    from plyer import notification
    NOTIFICATION_AVAILABLE = True
except ImportError:
    NOTIFICATION_AVAILABLE = False

# Playwright 관련 import (선택적)
try:
    from playwright.sync_api import sync_playwright
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PLAYWRIGHT_AVAILABLE = False

class WebCrawlerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("웹 크롤러")
        self.root.geometry("800x600")
        
        # 메인 프레임
        main_frame = ttk.Frame(root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # URL 입력 섹션
        url_frame = ttk.LabelFrame(main_frame, text="크롤링 설정", padding="5")
        url_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        ttk.Label(url_frame, text="URL (엔터키로 시작):").grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        self.url_var = tk.StringVar(value="https://example.com")
        self.url_entry = ttk.Entry(url_frame, textvariable=self.url_var, width=60)
        self.url_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 5))
        
        # URL 입력창에 키보드 단축키 바인딩 추가
        self.setup_url_entry_bindings()
        
        self.crawl_button = ttk.Button(url_frame, text="크롤링 시작", command=self.start_crawling)
        self.crawl_button.grid(row=0, column=2, padx=(5, 0))
        
        self.stop_button = ttk.Button(url_frame, text="중지", command=self.stop_crawling, state='disabled')
        self.stop_button.grid(row=0, column=3, padx=(5, 0))
        
        # 크롤링 옵션
        options_frame = ttk.LabelFrame(main_frame, text="크롤링 옵션", padding="5")
        options_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(5, 0))
        
        # 기본 추출 옵션
        extract_frame = ttk.Frame(options_frame)
        extract_frame.grid(row=0, column=0, columnspan=4, sticky=(tk.W, tk.E), pady=(0, 5))
        
        self.extract_links = tk.BooleanVar(value=True)
        ttk.Checkbutton(extract_frame, text="링크 추출", variable=self.extract_links).grid(row=0, column=0, sticky=tk.W)
        
        self.extract_images = tk.BooleanVar(value=True)
        ttk.Checkbutton(extract_frame, text="이미지 추출", variable=self.extract_images).grid(row=0, column=1, sticky=tk.W, padx=(10, 0))
        
        self.extract_text = tk.BooleanVar(value=True)
        ttk.Checkbutton(extract_frame, text="텍스트 추출", variable=self.extract_text).grid(row=0, column=2, sticky=tk.W, padx=(10, 0))
        
        # 고급 옵션
        advanced_frame = ttk.Frame(options_frame)
        advanced_frame.grid(row=1, column=0, columnspan=4, sticky=(tk.W, tk.E), pady=(5, 0))
        
        # 브라우저 모드
        ttk.Label(advanced_frame, text="브라우저 모드:").grid(row=0, column=0, sticky=tk.W)
        self.browser_mode = tk.StringVar(value="requests")
        
        # 사용 가능한 브라우저 모드 리스트
        browser_modes = ["requests", "selenium"]
        if PLAYWRIGHT_AVAILABLE:
            browser_modes.append("playwright")
        
        browser_combo = ttk.Combobox(advanced_frame, textvariable=self.browser_mode, values=browser_modes, width=12)
        browser_combo.grid(row=0, column=1, sticky=tk.W, padx=(5, 10))
        
        # 페이지 수
        ttk.Label(advanced_frame, text="페이지 수:").grid(row=0, column=2, sticky=tk.W)
        self.max_pages = tk.StringVar(value="1")
        ttk.Entry(advanced_frame, textvariable=self.max_pages, width=5).grid(row=0, column=3, sticky=tk.W, padx=(5, 10))
        
        # 크롤링 간격
        ttk.Label(advanced_frame, text="간격(초):").grid(row=0, column=4, sticky=tk.W)
        self.crawl_delay = tk.StringVar(value="1")
        ttk.Entry(advanced_frame, textvariable=self.crawl_delay, width=5).grid(row=0, column=5, sticky=tk.W, padx=(5, 10))
        
        # 사이트별 설정
        site_frame = ttk.Frame(options_frame)
        site_frame.grid(row=2, column=0, columnspan=4, sticky=(tk.W, tk.E), pady=(5, 0))
        
        ttk.Label(site_frame, text="사이트 종류:").grid(row=0, column=0, sticky=tk.W)
        self.site_type = tk.StringVar(value="일반")
        site_combo = ttk.Combobox(site_frame, textvariable=self.site_type, 
                                values=["일반", "네이버 쇼핑", "인스타그램", "부동산"], width=12)
        site_combo.grid(row=0, column=1, sticky=tk.W, padx=(5, 10))
        
        # 추출할 데이터 선택
        data_frame = ttk.Frame(options_frame)
        data_frame.grid(row=3, column=0, columnspan=4, sticky=(tk.W, tk.E), pady=(5, 0))
        
        ttk.Label(data_frame, text="추출 데이터:").grid(row=0, column=0, sticky=tk.W)
        self.extract_title = tk.BooleanVar(value=True)
        ttk.Checkbutton(data_frame, text="제목", variable=self.extract_title).grid(row=0, column=1, sticky=tk.W, padx=(5, 0))
        
        self.extract_price = tk.BooleanVar(value=True)
        ttk.Checkbutton(data_frame, text="가격", variable=self.extract_price).grid(row=0, column=2, sticky=tk.W, padx=(5, 0))
        
        self.extract_date = tk.BooleanVar(value=True)
        ttk.Checkbutton(data_frame, text="날짜", variable=self.extract_date).grid(row=0, column=3, sticky=tk.W, padx=(5, 0))
        
        self.extract_description = tk.BooleanVar(value=False)
        ttk.Checkbutton(data_frame, text="설명", variable=self.extract_description).grid(row=0, column=4, sticky=tk.W, padx=(5, 0))
        
        self.extract_images = tk.BooleanVar(value=False)
        ttk.Checkbutton(data_frame, text="이미지", variable=self.extract_images).grid(row=0, column=5, sticky=tk.W, padx=(5, 0))
        
        # 진행상황 표시
        self.progress = ttk.Progressbar(main_frame, mode='indeterminate')
        self.progress.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # 엑셀 저장 버튼
        export_frame = ttk.Frame(main_frame)
        export_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        self.export_button = ttk.Button(export_frame, text="엑셀로 저장", command=self.export_to_excel, state='disabled')
        self.export_button.pack(side=tk.RIGHT)
        
        ttk.Label(export_frame, text="크롤링이 완료되면 결과를 엑셀 파일로 저장할 수 있습니다.").pack(side=tk.LEFT)
        
        # 결과 표시 영역
        result_frame = ttk.LabelFrame(main_frame, text="크롤링 결과", padding="5")
        result_frame.grid(row=4, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        
        # 탭 위젯
        self.notebook = ttk.Notebook(result_frame)
        self.notebook.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 페이지 정보 탭
        self.info_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.info_frame, text="페이지 정보")
        
        self.info_text = scrolledtext.ScrolledText(self.info_frame, height=8, wrap=tk.WORD)
        self.info_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 링크 탭
        self.links_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.links_frame, text="링크")
        
        self.links_text = scrolledtext.ScrolledText(self.links_frame, height=8, wrap=tk.WORD)
        self.links_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 이미지 탭
        self.images_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.images_frame, text="이미지")
        
        self.images_text = scrolledtext.ScrolledText(self.images_frame, height=8, wrap=tk.WORD)
        self.images_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 텍스트 탭
        self.content_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.content_frame, text="텍스트 내용")
        
        self.content_text = scrolledtext.ScrolledText(self.content_frame, height=8, wrap=tk.WORD)
        self.content_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 결과 테이블 탭
        self.table_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.table_frame, text="결과 테이블")
        
        # 테이블 뷰 생성
        self.create_table_view()
        
        # 사이트 추천 탭
        self.recommend_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.recommend_frame, text="사이트 추천")
        
        self.recommend_text = scrolledtext.ScrolledText(self.recommend_frame, height=8, wrap=tk.WORD)
        self.recommend_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 기술스택 설명 탭
        self.tech_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.tech_frame, text="기술스택 가이드")
        
        self.tech_text = scrolledtext.ScrolledText(self.tech_frame, height=8, wrap=tk.WORD)
        self.tech_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 알림 설정 탭
        self.alert_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.alert_frame, text="알림 설정")
        self.setup_alert_tab()
        
        # 데이터 분석 탭
        self.analysis_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.analysis_frame, text="데이터 분석")
        self.setup_analysis_tab()
        
        # 스케줄링 탭
        self.schedule_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.schedule_frame, text="스케줄링")
        self.setup_schedule_tab()
        
        # 상태 표시줄
        self.status_var = tk.StringVar(value="준비")
        status_bar = ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN)
        status_bar.grid(row=5, column=0, columnspan=2, sticky=(tk.W, tk.E))
        
        # 그리드 가중치 설정
        root.columnconfigure(0, weight=1)
        root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(4, weight=1)
        result_frame.columnconfigure(0, weight=1)
        result_frame.rowconfigure(0, weight=1)
        
        # 각 탭의 그리드 설정
        for frame in [self.info_frame, self.links_frame, self.images_frame, self.content_frame, 
                      self.table_frame, self.recommend_frame, self.tech_frame, self.alert_frame, 
                      self.analysis_frame, self.schedule_frame]:
            frame.columnconfigure(0, weight=1)
            frame.rowconfigure(0, weight=1)
        
        url_frame.columnconfigure(1, weight=1)
        
        # 크롤링 상태 및 데이터 저장용 변수
        self.is_crawling = False
        self.crawled_data = []
        self.driver = None
        self.playwright = None
        self.browser = None
        self.page = None
        self.current_page = 1
        self.retry_count = 0
        self.max_retries = 3
        
        # 알림 및 분석 관련 변수
        self.alert_settings = {
            'email_enabled': False,
            'notification_enabled': True,
            'price_threshold': 0,
            'keywords': []
        }
        self.historical_data = []
        self.monitoring_active = False
        
        # 스케줄링 관련 변수
        self.scheduler = BackgroundScheduler()
        self.scheduler.start()
        self.scheduled_jobs = []
        
        # 중단점 재시작 관련 변수
        self.checkpoint_file = "crawling_checkpoint.pkl"
        self.current_task = None
        self.task_progress = {
            'total_pages': 0,
            'completed_pages': 0,
            'failed_pages': [],
            'current_url': '',
            'settings': {}
        }
        
        # 탭 내용 초기화
        self.load_static_content()
    
    def setup_alert_tab(self):
        """알림 설정 탭을 구성합니다."""
        # 스크롤 가능한 프레임
        canvas = tk.Canvas(self.alert_frame)
        scrollbar = ttk.Scrollbar(self.alert_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # 알림 설정 섹션
        alert_config = ttk.LabelFrame(scrollable_frame, text="알림 설정", padding="10")
        alert_config.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10), padx=10)
        
        # 데스크탑 알림 설정
        self.notification_enabled = tk.BooleanVar(value=True)
        ttk.Checkbutton(alert_config, text="데스크탑 알림 사용", 
                       variable=self.notification_enabled).grid(row=0, column=0, sticky=tk.W, pady=2)
        
        # 이메일 알림 설정
        self.email_enabled = tk.BooleanVar(value=False)
        ttk.Checkbutton(alert_config, text="이메일 알림 사용", 
                       variable=self.email_enabled).grid(row=1, column=0, sticky=tk.W, pady=2)
        
        # 이메일 설정 프레임
        email_frame = ttk.LabelFrame(alert_config, text="이메일 설정", padding="5")
        email_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
        ttk.Label(email_frame, text="발신 이메일:").grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        self.sender_email = tk.StringVar()
        ttk.Entry(email_frame, textvariable=self.sender_email, width=30).grid(row=0, column=1, sticky=(tk.W, tk.E))
        
        ttk.Label(email_frame, text="앱 비밀번호:").grid(row=1, column=0, sticky=tk.W, padx=(0, 5))
        self.sender_password = tk.StringVar()
        ttk.Entry(email_frame, textvariable=self.sender_password, show="*", width=30).grid(row=1, column=1, sticky=(tk.W, tk.E))
        
        ttk.Label(email_frame, text="수신 이메일:").grid(row=2, column=0, sticky=tk.W, padx=(0, 5))
        self.receiver_email = tk.StringVar()
        ttk.Entry(email_frame, textvariable=self.receiver_email, width=30).grid(row=2, column=1, sticky=(tk.W, tk.E))
        
        # 가격 알림 설정
        price_frame = ttk.LabelFrame(scrollable_frame, text="가격 알림 설정", padding="10")
        price_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10), padx=10)
        
        ttk.Label(price_frame, text="가격 변동 임계값:").grid(row=0, column=0, sticky=tk.W)
        self.price_threshold = tk.StringVar(value="10")
        ttk.Entry(price_frame, textvariable=self.price_threshold, width=10).grid(row=0, column=1, sticky=tk.W, padx=(5, 5))
        ttk.Label(price_frame, text="%").grid(row=0, column=2, sticky=tk.W)
        
        # 키워드 알림 설정
        keyword_frame = ttk.LabelFrame(scrollable_frame, text="키워드 알림 설정", padding="10")
        keyword_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 10), padx=10)
        
        ttk.Label(keyword_frame, text="모니터링 키워드:").grid(row=0, column=0, sticky=tk.W)
        self.keyword_entry = tk.StringVar()
        keyword_entry_widget = ttk.Entry(keyword_frame, textvariable=self.keyword_entry, width=30)
        keyword_entry_widget.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(5, 5))
        
        ttk.Button(keyword_frame, text="추가", command=self.add_keyword).grid(row=0, column=2, padx=(5, 0))
        
        # 키워드 목록
        self.keyword_listbox = tk.Listbox(keyword_frame, height=4)
        self.keyword_listbox.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(5, 0))
        
        ttk.Button(keyword_frame, text="키워드 삭제", command=self.remove_keyword).grid(row=2, column=0, pady=(5, 0))
        
        # 모니터링 제어 버튼
        control_frame = ttk.LabelFrame(scrollable_frame, text="모니터링 제어", padding="10")
        control_frame.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=(0, 10), padx=10)
        
        self.monitor_button = ttk.Button(control_frame, text="모니터링 시작", command=self.toggle_monitoring)
        self.monitor_button.grid(row=0, column=0, padx=(0, 5))
        
        ttk.Button(control_frame, text="알림 테스트", command=self.test_notification).grid(row=0, column=1, padx=(5, 0))
        
        # 스크롤바 설정
        canvas.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
    def setup_analysis_tab(self):
        """데이터 분석 탭을 구성합니다."""
        # 메인 프레임
        main_analysis_frame = ttk.Frame(self.analysis_frame)
        main_analysis_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=10, pady=10)
        
        # 분석 옵션 프레임
        options_frame = ttk.LabelFrame(main_analysis_frame, text="분석 옵션", padding="10")
        options_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # 차트 타입 선택
        ttk.Label(options_frame, text="차트 타입:").grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        self.chart_type = tk.StringVar(value="가격 추이")
        chart_combo = ttk.Combobox(options_frame, textvariable=self.chart_type, 
                                 values=["가격 추이", "키워드 빈도", "사이트별 통계", "시간대별 분포"])
        chart_combo.grid(row=0, column=1, sticky=tk.W, padx=(0, 10))
        
        # 분석 버튼
        ttk.Button(options_frame, text="차트 생성", command=self.generate_chart).grid(row=0, column=2, padx=(5, 0))
        ttk.Button(options_frame, text="리포트 생성", command=self.generate_report).grid(row=0, column=3, padx=(5, 0))
        
        # 차트 표시 영역
        self.chart_frame = ttk.LabelFrame(main_analysis_frame, text="차트", padding="5")
        self.chart_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        
        # 통계 정보 표시 영역
        stats_frame = ttk.LabelFrame(main_analysis_frame, text="통계 정보", padding="5")
        stats_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        self.stats_text = scrolledtext.ScrolledText(stats_frame, height=6, wrap=tk.WORD)
        self.stats_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 차트 사용 불가 시 안내
        if not CHART_AVAILABLE:
            ttk.Label(self.chart_frame, text="차트 기능을 사용하려면 matplotlib 라이브러리를 설치해주세요.\n'pip install matplotlib seaborn'").grid(row=0, column=0, pady=20)
    
    def setup_schedule_tab(self):
        """스케줄링 탭을 구성합니다."""
        # 스크롤 가능한 프레임
        canvas = tk.Canvas(self.schedule_frame)
        scrollbar = ttk.Scrollbar(self.schedule_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # 스케줄 설정 섹션
        schedule_config = ttk.LabelFrame(scrollable_frame, text="스케줄 설정", padding="10")
        schedule_config.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10), padx=10)
        
        # 스케줄 타입 선택
        ttk.Label(schedule_config, text="실행 주기:").grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        self.schedule_type = tk.StringVar(value="일회성")
        schedule_combo = ttk.Combobox(schedule_config, textvariable=self.schedule_type, 
                                    values=["일회성", "매일", "매주", "매월"], width=15)
        schedule_combo.grid(row=0, column=1, sticky=tk.W, padx=(0, 10))
        
        # 시간 설정
        ttk.Label(schedule_config, text="실행 시간:").grid(row=1, column=0, sticky=tk.W, padx=(0, 5))
        time_frame = ttk.Frame(schedule_config)
        time_frame.grid(row=1, column=1, sticky=tk.W, pady=5)
        
        self.schedule_hour = tk.StringVar(value="09")
        ttk.Spinbox(time_frame, from_=0, to=23, textvariable=self.schedule_hour, width=5, format="%02.0f").grid(row=0, column=0)
        ttk.Label(time_frame, text=":").grid(row=0, column=1)
        self.schedule_minute = tk.StringVar(value="00")
        ttk.Spinbox(time_frame, from_=0, to=59, textvariable=self.schedule_minute, width=5, format="%02.0f").grid(row=0, column=2)
        
        # 요일 선택 (주간 스케줄용)
        ttk.Label(schedule_config, text="요일 (매주):").grid(row=2, column=0, sticky=tk.W, padx=(0, 5))
        self.schedule_weekday = tk.StringVar(value="월요일")
        weekday_combo = ttk.Combobox(schedule_config, textvariable=self.schedule_weekday,
                                   values=["월요일", "화요일", "수요일", "목요일", "금요일", "토요일", "일요일"], width=15)
        weekday_combo.grid(row=2, column=1, sticky=tk.W)
        
        # 일자 설정 (월간 스케줄용)
        ttk.Label(schedule_config, text="일자 (매월):").grid(row=3, column=0, sticky=tk.W, padx=(0, 5))
        self.schedule_day = tk.StringVar(value="1")
        ttk.Spinbox(schedule_config, from_=1, to=28, textvariable=self.schedule_day, width=5).grid(row=3, column=1, sticky=tk.W)
        
        # 이메일 결과 전송 설정
        self.email_results = tk.BooleanVar(value=True)
        ttk.Checkbutton(schedule_config, text="결과를 이메일로 전송", 
                       variable=self.email_results).grid(row=4, column=0, columnspan=2, sticky=tk.W, pady=5)
        
        # 스케줄 제어 버튼
        control_frame = ttk.Frame(schedule_config)
        control_frame.grid(row=5, column=0, columnspan=2, pady=10)
        
        ttk.Button(control_frame, text="스케줄 추가", command=self.add_schedule).grid(row=0, column=0, padx=(0, 5))
        ttk.Button(control_frame, text="스케줄 삭제", command=self.remove_schedule).grid(row=0, column=1, padx=(5, 0))
        
        # 활성 스케줄 목록
        schedule_list_frame = ttk.LabelFrame(scrollable_frame, text="활성 스케줄", padding="10")
        schedule_list_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10), padx=10)
        
        # 스케줄 목록 트리뷰
        columns = ('ID', 'Type', 'Time', 'Next Run', 'Status')
        self.schedule_tree = ttk.Treeview(schedule_list_frame, columns=columns, show='headings', height=6)
        
        self.schedule_tree.heading('ID', text='ID')
        self.schedule_tree.heading('Type', text='타입')
        self.schedule_tree.heading('Time', text='시간')
        self.schedule_tree.heading('Next Run', text='다음 실행')
        self.schedule_tree.heading('Status', text='상태')
        
        self.schedule_tree.column('ID', width=50)
        self.schedule_tree.column('Type', width=80)
        self.schedule_tree.column('Time', width=100)
        self.schedule_tree.column('Next Run', width=150)
        self.schedule_tree.column('Status', width=80)
        
        self.schedule_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 복구 기능 섹션
        recovery_frame = ttk.LabelFrame(scrollable_frame, text="복구 기능", padding="10")
        recovery_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 10), padx=10)
        
        # 체크포인트 상태
        self.checkpoint_status = tk.StringVar(value="체크포인트 없음")
        ttk.Label(recovery_frame, textvariable=self.checkpoint_status).grid(row=0, column=0, columnspan=2, pady=5)
        
        # 복구 버튼
        recovery_control = ttk.Frame(recovery_frame)
        recovery_control.grid(row=1, column=0, columnspan=2, pady=5)
        
        ttk.Button(recovery_control, text="마지막 작업 복구", command=self.restore_last_task).grid(row=0, column=0, padx=(0, 5))
        ttk.Button(recovery_control, text="체크포인트 삭제", command=self.clear_checkpoint).grid(row=0, column=1, padx=(5, 0))
        
        # 자동 저장 설정
        self.auto_save = tk.BooleanVar(value=True)
        ttk.Checkbutton(recovery_frame, text="자동 체크포인트 저장", 
                       variable=self.auto_save).grid(row=2, column=0, columnspan=2, sticky=tk.W, pady=5)
        
        # 스크롤바 설정
        canvas.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        # 체크포인트 상태 업데이트
        self.update_checkpoint_status()
    
    def setup_url_entry_bindings(self):
        """URL 입력창에 키보드 단축키 바인딩을 설정합니다."""
        import platform
        
        # 운영체제 감지
        system = platform.system()
        
        if system == "Darwin":  # macOS
            # Cmd+A: 전체 선택
            self.url_entry.bind('<Command-a>', self.select_all_url)
            # Cmd+C: 복사
            self.url_entry.bind('<Command-c>', self.copy_url)
            # Cmd+V: 붙여넣기
            self.url_entry.bind('<Command-v>', self.paste_url)
            # Cmd+X: 잘라내기
            self.url_entry.bind('<Command-x>', self.cut_url)
            # Cmd+Z: 실행취소
            self.url_entry.bind('<Command-z>', self.undo_url)
        else:  # Windows, Linux
            # Ctrl+A: 전체 선택
            self.url_entry.bind('<Control-a>', self.select_all_url)
            # Ctrl+C: 복사
            self.url_entry.bind('<Control-c>', self.copy_url)
            # Ctrl+V: 붙여넣기
            self.url_entry.bind('<Control-v>', self.paste_url)
            # Ctrl+X: 잘라내기
            self.url_entry.bind('<Control-x>', self.cut_url)
            # Ctrl+Z: 실행취소
            self.url_entry.bind('<Control-z>', self.undo_url)
        
        # 우클릭 컨텍스트 메뉴 추가
        self.url_entry.bind('<Button-2>', self.show_context_menu)  # macOS
        self.url_entry.bind('<Button-3>', self.show_context_menu)  # Windows/Linux
        
        # 포커스 시 전체 선택 (선택사항)
        self.url_entry.bind('<FocusIn>', self.on_url_focus)
        
        # 엔터키로 크롤링 시작
        self.url_entry.bind('<Return>', self.on_url_enter)
        self.url_entry.bind('<KP_Enter>', self.on_url_enter)  # 숫자패드 엔터키도 지원
    
    def select_all_url(self, event=None):
        """URL 입력창의 모든 텍스트를 선택합니다."""
        self.url_entry.select_range(0, tk.END)
        return 'break'
    
    def copy_url(self, event=None):
        """선택된 텍스트를 클립보드로 복사합니다."""
        try:
            if self.url_entry.selection_present():
                text = self.url_entry.selection_get()
                self.root.clipboard_clear()
                self.root.clipboard_append(text)
        except tk.TclError:
            pass
        return 'break'
    
    def paste_url(self, event=None):
        """클립보드의 내용을 현재 커서 위치에 붙여넣습니다."""
        try:
            clipboard_text = self.root.clipboard_get()
            # 현재 선택된 텍스트가 있으면 대체, 없으면 커서 위치에 삽입
            if self.url_entry.selection_present():
                self.url_entry.delete(tk.SEL_FIRST, tk.SEL_LAST)
            self.url_entry.insert(tk.INSERT, clipboard_text)
        except tk.TclError:
            pass
        return 'break'
    
    def cut_url(self, event=None):
        """선택된 텍스트를 잘라내어 클립보드로 복사합니다."""
        try:
            if self.url_entry.selection_present():
                text = self.url_entry.selection_get()
                self.root.clipboard_clear()
                self.root.clipboard_append(text)
                self.url_entry.delete(tk.SEL_FIRST, tk.SEL_LAST)
        except tk.TclError:
            pass
        return 'break'
    
    def undo_url(self, event=None):
        """마지막 작업을 실행취소합니다."""
        try:
            self.url_entry.edit_undo()
        except tk.TclError:
            pass
        return 'break'
    
    def show_context_menu(self, event):
        """우클릭 컨텍스트 메뉴를 표시합니다."""
        # 컨텍스트 메뉴 생성
        context_menu = tk.Menu(self.root, tearoff=0)
        context_menu.add_command(label="잘라내기", command=lambda: self.cut_url())
        context_menu.add_command(label="복사", command=lambda: self.copy_url())
        context_menu.add_command(label="붙여넣기", command=lambda: self.paste_url())
        context_menu.add_separator()
        context_menu.add_command(label="전체 선택", command=lambda: self.select_all_url())
        context_menu.add_separator()
        context_menu.add_command(label="실행취소", command=lambda: self.undo_url())
        
        try:
            context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            context_menu.grab_release()
    
    def on_url_focus(self, event=None):
        """URL 입력창이 포커스를 받을 때 호출됩니다."""
        # 기본 URL이 있을 때만 전체 선택 (선택사항)
        if self.url_var.get() == "https://example.com":
            self.root.after(50, self.select_all_url)  # 약간의 지연 후 선택
    
    def on_url_enter(self, event=None):
        """URL 입력창에서 엔터키를 누를 때 호출됩니다."""
        # 크롤링이 진행 중이 아니고 크롤링 버튼이 활성화되어 있을 때만 실행
        if not self.is_crawling and self.crawl_button['state'] != 'disabled':
            # URL이 비어있지 않은지 확인
            url = self.url_var.get().strip()
            if url:
                # 크롤링 시작 버튼 클릭과 동일한 동작
                self.start_crawling()
                return 'break'  # 이벤트 전파 중단
        return None
    
    def create_table_view(self):
        """결과 테이블 뷰를 생성합니다."""
        # 테이블 프레임
        table_container = ttk.Frame(self.table_frame)
        table_container.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 테이블 (Treeview) 생성
        columns = ('Type', 'Title', 'URL', 'Description')
        self.tree = ttk.Treeview(table_container, columns=columns, show='headings', height=15)
        
        # 컬럼 헤더 설정
        self.tree.heading('Type', text='타입')
        self.tree.heading('Title', text='제목/텍스트')
        self.tree.heading('URL', text='URL')
        self.tree.heading('Description', text='설명')
        
        # 컬럼 너비 설정
        self.tree.column('Type', width=80, minwidth=60)
        self.tree.column('Title', width=200, minwidth=150)
        self.tree.column('URL', width=300, minwidth=200)
        self.tree.column('Description', width=150, minwidth=100)
        
        # 스크롤바 추가
        v_scrollbar = ttk.Scrollbar(table_container, orient=tk.VERTICAL, command=self.tree.yview)
        h_scrollbar = ttk.Scrollbar(table_container, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # 위젯 배치
        self.tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        v_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        h_scrollbar.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
        # 그리드 가중치 설정
        table_container.columnconfigure(0, weight=1)
        table_container.rowconfigure(0, weight=1)
    
    def stop_crawling(self):
        """크롤링을 중지합니다."""
        self.is_crawling = False
        self.status_var.set("크롤링 중지됨")
        
        # Selenium 드라이버 정리
        if self.driver:
            try:
                self.driver.quit()
            except:
                pass
            self.driver = None
        
        # Playwright 정리
        if self.page:
            try:
                self.page.close()
            except:
                pass
            self.page = None
        
        if self.browser:
            try:
                self.browser.close()
            except:
                pass
            self.browser = None
        
        if self.playwright:
            try:
                self.playwright.stop()
            except:
                pass
            self.playwright = None
    
    def start_crawling(self, scheduled=False):
        """백그라운드에서 크롤링을 시작합니다."""
        url = self.url_var.get().strip()
        if not url:
            if not scheduled:
                messagebox.showerror("오류", "URL을 입력해주세요.")
            return
        
        if not url.startswith(('http://', 'https://')):
            url = 'https://' + url
            self.url_var.set(url)
        
        # 현재 작업 정보 저장
        self.current_task = {
            'url': url,
            'browser_mode': self.browser_mode.get(),
            'max_pages': int(self.max_pages.get()) if self.max_pages.get().isdigit() else 1,
            'crawl_delay': float(self.crawl_delay.get()) if self.crawl_delay.get().replace('.', '').isdigit() else 1,
            'site_type': self.site_type.get(),
            'started_at': datetime.now(),
            'scheduled': scheduled
        }
        
        # 진행 상황 초기화
        self.task_progress = {
            'total_pages': self.current_task['max_pages'],
            'completed_pages': 0,
            'failed_pages': [],
            'current_url': url,
            'settings': self.current_task.copy()
        }
        
        # UI 상태 변경 (스케줄된 작업이 아닐 때만)
        if not scheduled:
            self.crawl_button.config(state='disabled')
            self.stop_button.config(state='normal')
            self.export_button.config(state='disabled')
            self.progress.start()
        
        self.status_var.set("크롤링 중...")
        self.is_crawling = True
        
        # 결과 영역 초기화
        if not scheduled:
            self.clear_results()
        self.crawled_data = []
        self.current_page = 1
        self.retry_count = 0
        
        # 체크포인트 저장
        if self.auto_save.get():
            self.save_checkpoint()
        
        # 백그라운드 스레드에서 크롤링 실행
        if self.browser_mode.get() == "selenium":
            thread = threading.Thread(target=self.crawl_with_selenium_checkpoint, args=(url, scheduled))
        elif self.browser_mode.get() == "playwright":
            thread = threading.Thread(target=self.crawl_with_playwright_checkpoint, args=(url, scheduled))
        else:
            thread = threading.Thread(target=self.crawl_website_checkpoint, args=(url, scheduled))
        thread.daemon = True
        thread.start()
    
    def clear_results(self):
        """결과 영역을 초기화합니다."""
        self.info_text.delete(1.0, tk.END)
        self.links_text.delete(1.0, tk.END)
        self.images_text.delete(1.0, tk.END)
        self.content_text.delete(1.0, tk.END)
        
        # 테이블 초기화
        for item in self.tree.get_children():
            self.tree.delete(item)
    
    def load_static_content(self):
        """정적인 탭 내용을 로드합니다."""
        # 사이트 추천 내용
        recommend_content = """
🌐 크롤링하기 쉬운 추천 사이트들

📰 뉴스 사이트
• https://news.ycombinator.com - 해커뉴스 (간단한 HTML 구조)
• https://quotes.toscrape.com - 크롤링 연습용 사이트
• https://books.toscrape.com - 도서 정보 연습 사이트

🏛️ 공공 데이터
• https://httpbin.org - HTTP 테스트용 사이트
• https://jsonplaceholder.typicode.com - JSON API 테스트
• https://reqres.in - API 테스트용

📊 연습용 사이트
• https://scrape.center - 크롤링 연습 전용
• https://webscraper.io/test-sites - 다양한 테스트 케이스
• https://example.com - 가장 기본적인 테스트

🛒 전자상거래 (주의 필요)
• 일부 전자상거래 사이트들 (robots.txt 준수 필수)
• 오픈 마켓 상품 정보 (이용약관 확인 필요)

⚠️ 주의사항
• robots.txt 파일을 항상 확인하세요
• 과도한 요청은 피하세요 (서버 부하 방지)
• 저작권이 있는 콘텐츠는 주의하세요
• 개인정보가 포함된 데이터는 수집하지 마세요
• 웹사이트의 이용약관을 반드시 확인하세요

🎯 초보자 추천 순서
1. quotes.toscrape.com (기본 연습)
2. books.toscrape.com (페이지네이션 연습)
3. news.ycombinator.com (실제 사이트 연습)
4. 원하는 사이트 (단계적 접근)
"""
        
        # 기술스택 가이드 내용
        tech_content = """
🌐 사이트 기술스택별 크롤링 난이도 가이드

📊 크롤링 난이도 분류

🟢 쉬움 (정적 사이트)
• 일반 HTML/CSS 사이트
• 서버 사이드 렌더링 (SSR)
• WordPress, Jekyll 등 정적 사이트 생성기

크롤링 방법:
→ requests + BeautifulSoup로 충분
→ 페이지 소스에 모든 데이터가 포함됨
→ 예: 블로그, 뉴스 사이트, 기업 홈페이지

🟡 보통 (하이브리드)
• jQuery 기반 사이트
• 일부 동적 로딩이 있는 사이트
• AJAX로 일부 콘텐츠 로드

크롤링 방법:
→ requests + BeautifulSoup (기본 콘텐츠)
→ API 엔드포인트 분석 후 직접 호출
→ 필요시 Selenium으로 동적 부분만 처리

🟠 어려움 (SPA - Single Page Application)
• React, Vue.js, Angular 기반
• 클라이언트 사이드 렌더링 (CSR)
• 대부분의 콘텐츠가 JavaScript로 생성

크롤링 방법:
→ Selenium 또는 Playwright 필수
→ 페이지 로딩 완료까지 대기 필요
→ 예: 모던 전자상거래, 소셜미디어

🔴 매우 어려움 (고도화된 SPA)
• Next.js, Nuxt.js (SSR + CSR 혼합)
• 복잡한 상태 관리 (Redux, Vuex)
• 지연 로딩, 무한 스크롤
• 강력한 봇 차단 시스템

크롤링 방법:
→ Playwright + 고급 기법 (권장)
→ 헤드리스 브라우저 + 프록시 로테이션
→ API 리버스 엔지니어링
→ 예: Netflix, Instagram, LinkedIn

🚀 브라우저 자동화 도구 비교

📊 Requests vs Selenium vs Playwright
┌─────────────┬────────────┬────────────┬──────────────┐
│   도구      │   속도     │   안정성   │   호환성     │
├─────────────┼────────────┼────────────┼──────────────┤
│ Requests    │ 매우 빠름  │ 높음       │ 정적 사이트  │
│ Selenium    │ 느림       │ 보통       │ 대부분 사이트│
│ Playwright  │ 빠름       │ 매우 높음  │ 모든 사이트  │
└─────────────┴────────────┴────────────┴──────────────┘

🎯 Playwright 장점:
• Selenium보다 2-3배 빠름
• 안정적인 요소 대기 및 선택
• 네트워크 인터셉션 지원
• 여러 브라우저 엔진 지원 (Chromium, Firefox, WebKit)
• 스크린샷 및 비디오 녹화 기능
• 향상된 디버깅 도구

🛠️ 기술스택별 상세 분석

📄 정적 HTML 사이트
특징: 서버에서 완전한 HTML 반환
도구: requests + BeautifulSoup
예시: 정부기관, 학교, 전통적 기업 사이트

🎯 WordPress/Drupal 사이트
특징: PHP 기반, 대부분 서버 렌더링
도구: requests + BeautifulSoup
주의: 플러그인에 따라 동적 요소 있을 수 있음

⚡ jQuery 기반 사이트
특징: 페이지 로드 후 일부 콘텐츠 추가 로드
도구: requests (기본) + Selenium (동적 부분)
전략: Network 탭에서 AJAX 요청 분석

🔥 React/Vue/Angular SPA
특징: 빈 HTML + JavaScript로 모든 콘텐츠 생성
도구: Selenium/Playwright 필수
전략: 
- componentDidMount 대기
- API 호출 직접 분석 (더 효율적)
- SSR 버전 찾기 (SEO용 페이지)

🏪 전자상거래 플랫폼별
• Shopify: Liquid 템플릿, 대부분 서버 렌더링
• WooCommerce: WordPress 기반, 비교적 쉬움
• Magento: 복잡한 구조, 일부 AJAX
• 커스텀 React: 매우 어려움

📱 모바일 앱 기반 웹
특징: 앱과 동일한 API 사용, 고도화된 보안
도구: API 리버스 엔지니어링 + requests
전략: 모바일 앱 패킷 분석

🛡️ 봇 차단 기술별 대응
• Cloudflare: User-Agent, 헤더 조작
• reCAPTCHA: 수동 해결 또는 우회
• Rate Limiting: 요청 간격 조절
• JavaScript Challenge: 헤드리스 브라우저 필수

🎯 실전 판별 방법

1️⃣ 페이지 소스 보기 (Ctrl+U)
→ 내용이 보이면: 정적 (쉬움)
→ 거의 비어있으면: SPA (어려움)

2️⃣ 개발자 도구 Network 탭
→ HTML 하나만: 정적
→ 많은 XHR/Fetch: 동적

3️⃣ JavaScript 비활성화 테스트
→ 내용이 그대로: 정적
→ 내용이 사라짐: 동적

4️⃣ URL 패턴 확인
→ /page/1, /category/tech: 전통적
→ /#/page/1, /app: SPA 가능성

💡 효율적인 접근 전략

1. 정적 분석 우선 (requests + BeautifulSoup)
2. 안 되면 API 분석 (Network 탭)
3. 마지막에 Selenium/Playwright 사용
4. 가능하면 모바일 버전이나 RSS 활용

⚠️ 주의사항
• robots.txt와 이용약관 확인 필수
• 과도한 요청 금지 (서버 부하)
• 개인정보 처리 시 법적 검토 필요
• 저작권 콘텐츠 주의
"""
        
        self.recommend_text.insert(tk.END, recommend_content)
        self.tech_text.insert(tk.END, tech_content)
        
        # 편집 불가능하게 설정
        self.recommend_text.config(state='disabled')
        self.tech_text.config(state='disabled')
    
    def crawl_website(self, url):
        """requests를 사용한 크롤링을 수행합니다."""
        try:
            max_pages = int(self.max_pages.get()) if self.max_pages.get().isdigit() else 1
            delay = float(self.crawl_delay.get()) if self.crawl_delay.get().replace('.', '').isdigit() else 1
            
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
            }
            
            for page in range(1, max_pages + 1):
                if not self.is_crawling:
                    break
                
                self.current_page = page
                self.root.after(0, lambda p=page: self.status_var.set(f"페이지 {p}/{max_pages} 크롤링 중..."))
                
                # 페이지별 URL 생성
                page_url = self.generate_page_url(url, page)
                
                # 재시도 로직
                success = False
                for retry in range(self.max_retries):
                    try:
                        response = requests.get(page_url, headers=headers, timeout=10)
                        response.raise_for_status()
                        
                        # 한글 인코딩 처리 개선
                        if response.encoding == 'ISO-8859-1':
                            # 잘못된 인코딩일 가능성이 높음
                            response.encoding = response.apparent_encoding
                        elif not response.encoding:
                            # 인코딩이 감지되지 않은 경우
                            response.encoding = 'utf-8'
                        
                        # 한글 사이트의 경우 추가 체크
                        if 'charset=' in response.headers.get('content-type', '').lower():
                            charset = response.headers.get('content-type', '').lower()
                            if 'euc-kr' in charset:
                                response.encoding = 'euc-kr'
                            elif 'utf-8' in charset:
                                response.encoding = 'utf-8'
                        else:
                            # apparent_encoding으로 자동 감지
                            response.encoding = response.apparent_encoding or 'utf-8'
                        
                        # BeautifulSoup으로 파싱
                        soup = BeautifulSoup(response.text, 'html.parser')
                        
                        # 결과 업데이트 (메인 스레드에서 실행)
                        self.root.after(0, self.update_results, page_url, response, soup)
                        
                        success = True
                        break
                        
                    except requests.exceptions.RequestException as e:
                        self.root.after(0, lambda r=retry+1: self.status_var.set(f"요청 실패, 재시도 {r}/{self.max_retries}"))
                        time.sleep(2)
                    except Exception as e:
                        self.root.after(0, lambda r=retry+1, err=str(e): self.status_var.set(f"오류 발생, 재시도 {r}/{self.max_retries}: {err[:50]}"))
                        time.sleep(2)
                
                if not success:
                    self.root.after(0, lambda p=page: self.status_var.set(f"페이지 {p} 크롤링 실패"))
                
                # 크롤링 간격
                if page < max_pages and self.is_crawling:
                    time.sleep(delay)
            
            self.root.after(0, self.finalize_crawling)
            
        except Exception as e:
            self.root.after(0, self.show_error, f"크롤링 오류: {str(e)}")
    
    def update_results(self, url, response, soup):
        """크롤링 결과를 UI에 업데이트합니다."""
        try:
            # 페이지 정보 - 한글 처리 개선
            title = soup.find('title')
            title_text = title.get_text(strip=True) if title else "제목 없음"
            
            # 제목 텍스트 정리
            if title_text:
                import unicodedata
                title_text = unicodedata.normalize('NFC', title_text)
                title_text = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', title_text)
                title_text = re.sub(r'\s+', ' ', title_text).strip()
            
            meta_desc = soup.find('meta', attrs={'name': 'description'})
            description = meta_desc.get('content', '설명 없음') if meta_desc else "설명 없음"
            
            # 설명 텍스트 정리
            if description and description != '설명 없음':
                description = unicodedata.normalize('NFC', description)
                description = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', description)
                description = re.sub(r'\s+', ' ', description).strip()
            
            info = f"제목: {title_text}\n"
            info += f"URL: {url}\n"
            info += f"상태 코드: {response.status_code}\n"
            info += f"콘텐츠 타입: {response.headers.get('content-type', '알 수 없음')}\n"
            info += f"콘텐츠 크기: {len(response.content)} bytes\n"
            info += f"설명: {description}\n"
            
            self.info_text.insert(tk.END, info)
            
            # 링크 추출
            if self.extract_links.get():
                links = soup.find_all('a', href=True)
                self.links_text.insert(tk.END, f"총 {len(links)}개의 링크를 발견했습니다:\n\n")
                
                for i, link in enumerate(links[:50], 1):  # 최대 50개만 표시
                    href = link['href']
                    text = link.get_text(strip=True)
                    
                    # 링크 텍스트 정리
                    if text:
                        text = unicodedata.normalize('NFC', text)
                        text = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', text)
                        text = re.sub(r'\s+', ' ', text).strip()
                        # 너무 긴 텍스트는 제한
                        if len(text) > 100:
                            text = text[:100] + "..."
                    else:
                        text = "(텍스트 없음)"
                    
                    full_url = urljoin(url, href)
                    self.links_text.insert(tk.END, f"{i}. {text}\n   URL: {full_url}\n\n")
                
                if len(links) > 50:
                    self.links_text.insert(tk.END, f"... 그 외 {len(links) - 50}개 링크가 더 있습니다.")
            
            # 이미지 추출
            if self.extract_images.get():
                images = soup.find_all('img', src=True)
                self.images_text.insert(tk.END, f"총 {len(images)}개의 이미지를 발견했습니다:\n\n")
                
                for i, img in enumerate(images[:30], 1):  # 최대 30개만 표시
                    src = img['src']
                    alt = img.get('alt', '대체 텍스트 없음')
                    full_url = urljoin(url, src)
                    self.images_text.insert(tk.END, f"{i}. {alt}\n   URL: {full_url}\n\n")
                
                if len(images) > 30:
                    self.images_text.insert(tk.END, f"... 그 외 {len(images) - 30}개 이미지가 더 있습니다.")
            
            # 텍스트 내용 추출
            if self.extract_text.get():
                # 스크립트와 스타일 태그 제거
                for script in soup(["script", "style"]):
                    script.decompose()
                
                text = soup.get_text()
                lines = (line.strip() for line in text.splitlines())
                chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
                text = ' '.join(chunk for chunk in chunks if chunk)
                
                # 한글 텍스트 정리
                import unicodedata
                # 유니코드 정규화 (한글 호환성 문자 처리)
                text = unicodedata.normalize('NFC', text)
                # 제어 문자 제거
                text = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', text)
                # 연속된 공백 정리
                text = re.sub(r'\s+', ' ', text).strip()
                
                # 텍스트 길이 제한
                if len(text) > 5000:
                    text = text[:5000] + "\n\n... (텍스트가 너무 길어 일부만 표시됩니다)"
                
                self.content_text.insert(tk.END, text)
            
            # 테이블에 데이터 추가
            self.populate_table(url, soup)
            
            self.status_var.set(f"크롤링 완료 - {title_text}")
            
        except Exception as e:
            self.show_error(f"결과 처리 오류: {str(e)}")
    
    def show_error(self, error_message):
        """오류 메시지를 표시합니다."""
        self.progress.stop()
        self.crawl_button.config(state='normal')
        self.stop_button.config(state='disabled')
        self.is_crawling = False
        self.status_var.set("오류 발생")
        messagebox.showerror("크롤링 오류", error_message)
    
    def populate_table(self, url, soup):
        """테이블에 크롤링 결과를 추가합니다."""
        try:
            # 페이지 기본 정보
            title = soup.find('title')
            title_text = title.get_text(strip=True) if title else "제목 없음"
            
            # 페이지 정보 추가
            page_data = {
                'type': '페이지',
                'title': title_text,
                'url': url,
                'description': '웹페이지 기본 정보'
            }
            self.tree.insert('', 'end', values=(page_data['type'], page_data['title'], 
                                              page_data['url'], page_data['description']))
            self.crawled_data.append(page_data)
            
            # 링크 정보 추가
            if self.extract_links.get():
                links = soup.find_all('a', href=True)
                for i, link in enumerate(links[:50]):  # 최대 50개
                    href = link['href']
                    text = link.get_text(strip=True)
                    
                    # 링크 텍스트 정리
                    if text:
                        import unicodedata
                        text = unicodedata.normalize('NFC', text)
                        text = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', text)
                        text = re.sub(r'\s+', ' ', text).strip()
                        # 길이 제한
                        if len(text) > 100:
                            text = text[:100] + "..."
                    else:
                        text = '(텍스트 없음)'
                    
                    full_url = urljoin(url, href)
                    
                    link_data = {
                        'type': '링크',
                        'title': text,
                        'url': full_url,
                        'description': f'링크 #{i+1}'
                    }
                    self.tree.insert('', 'end', values=(link_data['type'], link_data['title'], 
                                                      link_data['url'], link_data['description']))
                    self.crawled_data.append(link_data)
            
            # 이미지 정보 추가
            if self.extract_images.get():
                images = soup.find_all('img', src=True)
                for i, img in enumerate(images[:30]):  # 최대 30개
                    src = img['src']
                    alt = img.get('alt', '대체 텍스트 없음')[:100]
                    full_url = urljoin(url, src)
                    
                    img_data = {
                        'type': '이미지',
                        'title': alt,
                        'url': full_url,
                        'description': f'이미지 #{i+1}'
                    }
                    self.tree.insert('', 'end', values=(img_data['type'], img_data['title'], 
                                                      img_data['url'], img_data['description']))
                    self.crawled_data.append(img_data)
                    
        except Exception as e:
            print(f"테이블 채우기 오류: {str(e)}")
    
    def export_to_excel(self):
        """크롤링 결과를 엑셀 파일로 저장합니다."""
        if not self.crawled_data:
            messagebox.showwarning("경고", "저장할 데이터가 없습니다.")
            return
        
        try:
            # 파일 저장 대화상자 - 한글 파일명 지원
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"크롤링결과_{timestamp}.xlsx"  # 한글 파일명 사용
            
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")],
                initialvalue=default_filename,
                title="크롤링 결과 저장"
            )
            
            if not file_path:
                return
            
            # 한글 경로 처리를 위한 정규화
            file_path = os.path.normpath(file_path)
            
            # DataFrame 생성 - 한글 데이터 처리 개선
            processed_data = []
            for item in self.crawled_data:
                processed_item = {}
                for key, value in item.items():
                    # 텍스트 데이터 정리 및 인코딩 처리
                    if isinstance(value, str):
                        # 유니코드 정규화
                        value = unicodedata.normalize('NFC', value)
                        # 특수문자 및 제어문자 제거
                        value = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', value)
                        # 연속된 공백 정리
                        value = re.sub(r'\s+', ' ', value).strip()
                        # Excel 셀 길이 제한 (32767자)
                        if len(value) > 32000:
                            value = value[:32000] + "..."
                    processed_item[key] = value
                processed_data.append(processed_item)
            
            df = pd.DataFrame(processed_data)
            df.columns = ['타입', '제목/텍스트', 'URL', '설명']
            
            # 엑셀 파일로 저장 - 한글 지원 강화
            with pd.ExcelWriter(
                file_path, 
                engine='openpyxl'
            ) as writer:
                # 시트명을 영문으로 변경 (일부 Excel 버전 호환성)
                sheet_name = 'CrawlingResults'
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 워크시트 가져오기 및 서식 설정
                worksheet = writer.sheets[sheet_name]
                
                # 스타일 import를 try-except로 감싸서 오류 방지
                try:
                    from openpyxl.styles import Font, PatternFill, Alignment
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    
                    # 헤더 행 스타일 적용
                    for cell in worksheet[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                except ImportError:
                    # 스타일 라이브러리를 불러올 수 없는 경우 무시
                    pass
                
                # 컬럼 너비 자동 조정 - 한글 텍스트 고려
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if cell.value:
                                # 한글 텍스트의 실제 표시 길이 계산
                                text = str(cell.value)
                                # 한글은 2배, 영문은 1배로 계산
                                display_length = sum(2 if ord(char) > 127 else 1 for char in text)
                                if display_length > max_length:
                                    max_length = display_length
                        except:
                            pass
                    
                    # 최소 8, 최대 60으로 제한
                    adjusted_width = min(max(max_length * 0.8, 8), 60)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # 행 높이 자동 조정
                for row in worksheet.iter_rows():
                    max_lines = 1
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            lines = cell.value.count('\n') + 1
                            max_lines = max(max_lines, lines)
                    if max_lines > 1:
                        worksheet.row_dimensions[row[0].row].height = max_lines * 15
            
            # 저장 완료 메시지 - 파일 크기 정보 추가
            file_size = os.path.getsize(file_path)
            size_str = f"{file_size:,} bytes"
            if file_size > 1024:
                size_str = f"{file_size/1024:.1f} KB"
            if file_size > 1024*1024:
                size_str = f"{file_size/(1024*1024):.1f} MB"
            
            messagebox.showinfo(
                "저장 완료", 
                f"크롤링 결과가 성공적으로 저장되었습니다.\n\n"
                f"파일: {os.path.basename(file_path)}\n"
                f"위치: {os.path.dirname(file_path)}\n"
                f"크기: {size_str}\n"
                f"데이터 수: {len(self.crawled_data)}개"
            )
            
        except UnicodeEncodeError as e:
            messagebox.showerror("인코딩 오류", 
                f"파일 저장 중 인코딩 오류가 발생했습니다.\n"
                f"일부 특수문자가 포함되어 있을 수 있습니다.\n\n"
                f"오류 상세: {str(e)}")
        except PermissionError:
            messagebox.showerror("권한 오류", 
                f"파일을 저장할 권한이 없거나 파일이 다른 프로그램에서 사용 중입니다.\n"
                f"다른 위치에 저장하거나 파일을 닫고 다시 시도해주세요.")
        except Exception as e:
            messagebox.showerror("저장 오류", 
                f"파일 저장 중 오류가 발생했습니다.\n\n"
                f"오류 내용: {str(e)}\n\n"
                f"해결 방법:\n"
                f"1. 다른 파일명으로 저장해보세요\n"
                f"2. 다른 폴더에 저장해보세요\n"
                f"3. 프로그램을 관리자 권한으로 실행해보세요")
    
    def setup_selenium_driver(self):
        """Selenium WebDriver를 설정합니다."""
        try:
            chrome_options = Options()
            chrome_options.add_argument("--headless")  # 백그라운드 실행
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--window-size=1920,1080")
            chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])
            chrome_options.add_experimental_option('useAutomationExtension', False)
            chrome_options.add_argument("--disable-blink-features=AutomationControlled")
            
            # User-Agent 설정
            chrome_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
            
            service = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            
            # 자동화 감지 방지
            self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            
            return True
        except Exception as e:
            self.root.after(0, self.show_error, f"Selenium 설정 오류: {str(e)}")
            return False
    
    def setup_playwright_browser(self):
        """Playwright 브라우저를 설정합니다."""
        try:
            if not PLAYWRIGHT_AVAILABLE:
                self.root.after(0, self.show_error, "Playwright가 설치되지 않았습니다. pip install playwright 후 playwright install을 실행하세요.")
                return False
            
            self.playwright = sync_playwright().start()
            
            # 브라우저 선택 (Chromium 기본)
            self.browser = self.playwright.chromium.launch(
                headless=True,  # 백그라운드 실행
                args=[
                    '--no-sandbox',
                    '--disable-dev-shm-usage',
                    '--disable-blink-features=AutomationControlled'
                ]
            )
            
            # 컨텍스트 생성 (시크릿 모드와 유사)
            context = self.browser.new_context(
                user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
                viewport={'width': 1920, 'height': 1080}
            )
            
            # 페이지 생성
            self.page = context.new_page()
            
            # 자동화 감지 방지
            self.page.add_init_script("""
                Object.defineProperty(navigator, 'webdriver', {
                    get: () => undefined,
                });
            """)
            
            return True
        except Exception as e:
            self.root.after(0, self.show_error, f"Playwright 설정 오류: {str(e)}")
            return False
    
    def crawl_with_selenium(self, url):
        """Selenium을 사용한 크롤링을 수행합니다."""
        if not self.setup_selenium_driver():
            return
        
        try:
            max_pages = int(self.max_pages.get()) if self.max_pages.get().isdigit() else 1
            delay = float(self.crawl_delay.get()) if self.crawl_delay.get().replace('.', '').isdigit() else 1
            
            for page in range(1, max_pages + 1):
                if not self.is_crawling:
                    break
                
                self.current_page = page
                self.root.after(0, lambda p=page: self.status_var.set(f"페이지 {p}/{max_pages} 크롤링 중..."))
                
                # 페이지별 URL 생성
                page_url = self.generate_page_url(url, page)
                
                # 재시도 로직
                success = False
                for retry in range(self.max_retries):
                    try:
                        self.driver.get(page_url)
                        WebDriverWait(self.driver, 10).until(
                            EC.presence_of_element_located((By.TAG_NAME, "body"))
                        )
                        
                        # 사이트별 특화 크롤링
                        if self.site_type.get() == "네이버 쇼핑":
                            self.crawl_naver_shopping()
                        elif self.site_type.get() == "인스타그램":
                            self.crawl_instagram()
                        elif self.site_type.get() == "부동산":
                            self.crawl_real_estate()
                        else:
                            self.crawl_general_selenium()
                        
                        success = True
                        break
                        
                    except TimeoutException:
                        self.root.after(0, lambda r=retry+1: self.status_var.set(f"페이지 로딩 실패, 재시도 {r}/{self.max_retries}"))
                        time.sleep(2)
                    except Exception as e:
                        self.root.after(0, lambda r=retry+1, err=str(e): self.status_var.set(f"오류 발생, 재시도 {r}/{self.max_retries}: {err[:50]}"))
                        time.sleep(2)
                
                if not success:
                    self.root.after(0, lambda p=page: self.status_var.set(f"페이지 {p} 크롤링 실패"))
                
                # 크롤링 간격
                if page < max_pages and self.is_crawling:
                    time.sleep(delay)
            
            self.root.after(0, self.finalize_crawling)
            
        except Exception as e:
            self.root.after(0, self.show_error, f"Selenium 크롤링 오류: {str(e)}")
        finally:
            if self.driver:
                try:
                    self.driver.quit()
                except:
                    pass
                self.driver = None
    
    def generate_page_url(self, base_url, page):
        """페이지 번호에 따른 URL을 생성합니다."""
        if page == 1:
            return base_url
        
        # 일반적인 페이지네이션 패턴들
        if "?" in base_url:
            return f"{base_url}&page={page}"
        else:
            return f"{base_url}?page={page}"
    
    def crawl_naver_shopping(self):
        """네이버 쇼핑 크롤링"""
        try:
            # 상품 목록 요소 대기
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ".basicList_list__2YY7H, .product_list, .list_basis"))
            )
            
            # 상품 요소들 찾기
            products = self.driver.find_elements(By.CSS_SELECTOR, ".basicList_item__1MBN3, .product_item, .item")
            
            for product in products:
                if not self.is_crawling:
                    break
                
                try:
                    data = {'type': '네이버쇼핑'}
                    
                    # 제목 추출
                    if self.extract_title.get():
                        title_elem = product.find_element(By.CSS_SELECTOR, ".basicList_title__3P9Q7, .product_title, .item_title")
                        data['title'] = title_elem.text.strip()
                    
                    # 가격 추출
                    if self.extract_price.get():
                        price_elem = product.find_element(By.CSS_SELECTOR, ".price_num__2WUXn, .product_price, .item_price")
                        data['price'] = price_elem.text.strip()
                    
                    # 링크 추출
                    link_elem = product.find_element(By.CSS_SELECTOR, "a")
                    data['url'] = link_elem.get_attribute('href')
                    
                    # 이미지 추출
                    if self.extract_images.get():
                        img_elem = product.find_element(By.CSS_SELECTOR, "img")
                        data['image_url'] = img_elem.get_attribute('src')
                    
                    data['description'] = f"네이버쇼핑 상품 (페이지 {self.current_page})"
                    
                    self.crawled_data.append(data)
                    self.root.after(0, self.add_to_table, data)
                    
                except NoSuchElementException:
                    continue
                    
        except TimeoutException:
            self.root.after(0, lambda: self.status_var.set("네이버 쇼핑 요소를 찾을 수 없습니다"))
    
    def crawl_instagram(self):
        """인스타그램 크롤링 (제한적)"""
        try:
            # 인스타그램은 로그인이 필요하므로 기본적인 메타데이터만 추출
            posts = self.driver.find_elements(By.CSS_SELECTOR, "article, ._aagu, .v1Nh3")
            
            for i, post in enumerate(posts[:10]):  # 최대 10개만
                if not self.is_crawling:
                    break
                
                try:
                    data = {'type': '인스타그램'}
                    
                    # 기본 정보
                    data['title'] = f"Instagram Post {i+1}"
                    data['url'] = self.driver.current_url
                    data['description'] = f"인스타그램 게시물 (페이지 {self.current_page})"
                    
                    # 이미지 추출 시도
                    if self.extract_images.get():
                        try:
                            img_elem = post.find_element(By.CSS_SELECTOR, "img")
                            data['image_url'] = img_elem.get_attribute('src')
                        except:
                            pass
                    
                    self.crawled_data.append(data)
                    self.root.after(0, self.add_to_table, data)
                    
                except:
                    continue
                    
        except Exception as e:
            self.root.after(0, lambda err=str(e): self.status_var.set(f"인스타그램 크롤링 제한: {err[:50]}"))
    
    def crawl_real_estate(self):
        """부동산 사이트 크롤링"""
        try:
            # 부동산 매물 요소들 찾기
            properties = self.driver.find_elements(By.CSS_SELECTOR, ".item, .list-item, .property-item, .estate-item")
            
            for prop in properties:
                if not self.is_crawling:
                    break
                
                try:
                    data = {'type': '부동산'}
                    
                    # 제목 (주소/매물명)
                    if self.extract_title.get():
                        title_selectors = [".item-title", ".property-title", ".title", "h3", "h4"]
                        for selector in title_selectors:
                            try:
                                title_elem = prop.find_element(By.CSS_SELECTOR, selector)
                                data['title'] = title_elem.text.strip()
                                break
                            except:
                                continue
                    
                    # 가격 추출
                    if self.extract_price.get():
                        price_selectors = [".price", ".cost", ".amount", ".fee"]
                        for selector in price_selectors:
                            try:
                                price_elem = prop.find_element(By.CSS_SELECTOR, selector)
                                data['price'] = price_elem.text.strip()
                                break
                            except:
                                continue
                    
                    # 링크 추출
                    try:
                        link_elem = prop.find_element(By.CSS_SELECTOR, "a")
                        data['url'] = link_elem.get_attribute('href')
                    except:
                        data['url'] = self.driver.current_url
                    
                    data['description'] = f"부동산 매물 (페이지 {self.current_page})"
                    
                    self.crawled_data.append(data)
                    self.root.after(0, self.add_to_table, data)
                    
                except:
                    continue
                    
        except Exception as e:
            self.root.after(0, lambda err=str(e): self.status_var.set(f"부동산 크롤링 오류: {err[:50]}"))
    
    def crawl_general_selenium(self):
        """일반적인 Selenium 크롤링"""
        try:
            soup = BeautifulSoup(self.driver.page_source, 'html.parser')
            
            # 기존 populate_table 메서드 활용
            self.root.after(0, self.populate_table, self.driver.current_url, soup)
            
        except Exception as e:
            self.root.after(0, lambda err=str(e): self.status_var.set(f"일반 크롤링 오류: {err[:50]}"))
    
    def add_to_table(self, data):
        """테이블에 개별 데이터를 추가합니다."""
        try:
            self.tree.insert('', 'end', values=(
                data.get('type', ''),
                data.get('title', '')[:100],
                data.get('url', ''),
                data.get('description', '')
            ))
        except Exception as e:
            print(f"테이블 추가 오류: {str(e)}")
    
    def finalize_crawling(self):
        """크롤링 완료 처리"""
        self.progress.stop()
        self.crawl_button.config(state='normal')
        self.stop_button.config(state='disabled')
        self.export_button.config(state='normal')
        self.is_crawling = False
        self.status_var.set(f"크롤링 완료 - 총 {len(self.crawled_data)}개 데이터 수집")

    def crawl_with_playwright(self, url):
        """Playwright를 사용한 크롤링을 수행합니다."""
        try:
            if not self.setup_playwright_browser():
                return
            
            max_pages = int(self.max_pages.get()) if self.max_pages.get().isdigit() else 1
            delay = float(self.crawl_delay.get()) if self.crawl_delay.get().replace('.', '').isdigit() else 1
            
            for page in range(1, max_pages + 1):
                if not self.is_crawling:
                    break
                
                self.current_page = page
                self.root.after(0, lambda p=page: self.status_var.set(f"페이지 {p}/{max_pages} Playwright 크롤링 중..."))
                
                page_url = self.generate_page_url(url, page)
                
                success = False
                for retry in range(self.max_retries):
                    try:
                        # 페이지 로드
                        self.page.goto(page_url, wait_until='domcontentloaded', timeout=30000)
                        
                        # 네트워크 요청 완료까지 대기
                        self.page.wait_for_load_state('networkidle', timeout=10000)
                        
                        # 추가 대기 (JavaScript 실행 완료)
                        self.page.wait_for_timeout(2000)
                        
                        # 사이트별 특화 크롤링
                        domain = urlparse(page_url).netloc.lower()
                        if 'shopping.naver.com' in domain:
                            self.crawl_naver_shopping_playwright()
                        elif 'instagram.com' in domain:
                            self.crawl_instagram_playwright()
                        elif any(keyword in domain for keyword in ['zigbang', 'dabang', '부동산']):
                            self.crawl_real_estate_playwright()
                        else:
                            self.crawl_general_playwright()
                        
                        success = True
                        break
                        
                    except Exception as e:
                        self.root.after(0, lambda r=retry+1, err=str(e): self.status_var.set(f"Playwright 오류, 재시도 {r}/{self.max_retries}: {err[:30]}"))
                        time.sleep(2)
                
                if not success:
                    self.root.after(0, lambda p=page: self.status_var.set(f"페이지 {p} Playwright 크롤링 실패"))
                
                if page < max_pages and self.is_crawling:
                    time.sleep(delay)
            
            self.root.after(0, self.finalize_crawling)
            
        except Exception as e:
            self.root.after(0, self.show_error, f"Playwright 크롤링 오류: {str(e)}")
        finally:
            # Playwright 정리
            if self.page:
                try:
                    self.page.close()
                except:
                    pass
                self.page = None
            
            if self.browser:
                try:
                    self.browser.close()
                except:
                    pass
                self.browser = None
            
            if self.playwright:
                try:
                    self.playwright.stop()
                except:
                    pass
                self.playwright = None
    
    def crawl_general_playwright(self):
        """일반적인 Playwright 크롤링"""
        try:
            # 페이지 소스 가져오기
            content = self.page.content()
            soup = BeautifulSoup(content, 'html.parser')
            
            # 기존 populate_table 메서드 활용
            self.root.after(0, self.populate_table, self.page.url, soup)
            
        except Exception as e:
            self.root.after(0, lambda err=str(e): self.status_var.set(f"일반 Playwright 크롤링 오류: {err[:50]}"))
    
    def crawl_naver_shopping_playwright(self):
        """네이버 쇼핑 Playwright 크롤링"""
        try:
            # 상품 목록 대기
            self.page.wait_for_selector('.basicList_list_basis__uNBZx, .product_list, .goods_list', timeout=10000)
            
            # 상품 요소들 가져오기
            products = self.page.query_selector_all('.basicList_item__FxDgW, .product_item, .goods_item')
            
            for i, product in enumerate(products[:20]):  # 최대 20개
                if not self.is_crawling:
                    break
                
                try:
                    data = {'type': '네이버쇼핑'}
                    
                    # 상품명
                    if self.extract_title.get():
                        title_elem = product.query_selector('.basicList_title__3P9Q7, .product_title, .goods_name')
                        if title_elem:
                            data['title'] = title_elem.text_content().strip()
                        else:
                            data['title'] = f"상품 {i+1}"
                    
                    # 가격
                    if self.extract_price.get():
                        price_elem = product.query_selector('.price_num__2WUXn, .product_price, .price')
                        if price_elem:
                            data['price'] = price_elem.text_content().strip()
                    
                    # 상품 링크
                    link_elem = product.query_selector('a')
                    if link_elem:
                        data['url'] = link_elem.get_attribute('href')
                        if data['url'] and not data['url'].startswith('http'):
                            data['url'] = 'https://shopping.naver.com' + data['url']
                    else:
                        data['url'] = self.page.url
                    
                    # 설명
                    data['description'] = f"네이버 쇼핑 상품 (페이지 {self.current_page})"
                    
                    self.crawled_data.append(data)
                    self.root.after(0, self.add_to_table, data)
                    
                except Exception:
                    continue
                    
        except Exception as e:
            self.root.after(0, lambda err=str(e): self.status_var.set(f"네이버 쇼핑 Playwright 크롤링 오류: {err[:50]}"))
    
    def crawl_instagram_playwright(self):
        """인스타그램 Playwright 크롤링"""
        try:
            # 게시물 요소들 대기
            self.page.wait_for_selector('article, ._aagu', timeout=10000)
            
            posts = self.page.query_selector_all('article, ._aagu')
            
            for i, post in enumerate(posts[:10]):  # 최대 10개
                if not self.is_crawling:
                    break
                
                try:
                    data = {'type': '인스타그램'}
                    data['title'] = f"Instagram Post {i+1}"
                    data['url'] = self.page.url
                    data['description'] = f"인스타그램 게시물 (페이지 {self.current_page})"
                    
                    # 이미지 URL 추출
                    if self.extract_images.get():
                        img_elem = post.query_selector('img')
                        if img_elem:
                            data['image_url'] = img_elem.get_attribute('src')
                    
                    self.crawled_data.append(data)
                    self.root.after(0, self.add_to_table, data)
                    
                except Exception:
                    continue
                    
        except Exception as e:
            self.root.after(0, lambda err=str(e): self.status_var.set(f"인스타그램 Playwright 크롤링 오류: {err[:50]}"))
    
    def crawl_real_estate_playwright(self):
        """부동산 사이트 Playwright 크롤링"""
        try:
            # 매물 목록 대기
            self.page.wait_for_selector('.item, .property, .list-item', timeout=10000)
            
            properties = self.page.query_selector_all('.item, .property, .list-item')
            
            for i, prop in enumerate(properties[:15]):  # 최대 15개
                if not self.is_crawling:
                    break
                
                try:
                    data = {'type': '부동산'}
                    
                    # 제목 (주소/매물명)
                    if self.extract_title.get():
                        title_elem = prop.query_selector('.item-title, .property-title, .title, h3, h4')
                        if title_elem:
                            data['title'] = title_elem.text_content().strip()
                        else:
                            data['title'] = f"매물 {i+1}"
                    
                    # 가격
                    if self.extract_price.get():
                        price_elem = prop.query_selector('.price, .item-price, .property-price')
                        if price_elem:
                            data['price'] = price_elem.text_content().strip()
                    
                    # 매물 링크
                    link_elem = prop.query_selector('a')
                    if link_elem:
                        data['url'] = link_elem.get_attribute('href')
                        if data['url'] and not data['url'].startswith('http'):
                            base_url = f"{self.page.url.split('/')[0]}//{self.page.url.split('/')[2]}"
                            data['url'] = base_url + data['url']
                    else:
                        data['url'] = self.page.url
                    
                    # 설명
                    data['description'] = f"부동산 매물 (페이지 {self.current_page})"
                    
                    self.crawled_data.append(data)
                    self.root.after(0, self.add_to_table, data)
                    
                except Exception:
                    continue
                    
        except Exception as e:
            self.root.after(0, lambda err=str(e): self.status_var.set(f"부동산 Playwright 크롤링 오류: {err[:50]}"))

    # 알림 관련 메서드들
    def add_keyword(self):
        """키워드를 추가합니다."""
        keyword = self.keyword_entry.get().strip()
        if keyword and keyword not in self.keyword_listbox.get(0, tk.END):
            self.keyword_listbox.insert(tk.END, keyword)
            self.keyword_entry.set("")
    
    def remove_keyword(self):
        """선택된 키워드를 삭제합니다."""
        try:
            selection = self.keyword_listbox.curselection()
            if selection:
                self.keyword_listbox.delete(selection[0])
        except:
            pass
    
    def toggle_monitoring(self):
        """모니터링을 시작/중지합니다."""
        if self.monitoring_active:
            self.monitoring_active = False
            self.monitor_button.config(text="모니터링 시작")
            self.status_var.set("모니터링 중지됨")
        else:
            self.monitoring_active = True
            self.monitor_button.config(text="모니터링 중지")
            self.status_var.set("모니터링 시작됨")
            # 백그라운드에서 모니터링 실행
            threading.Thread(target=self.run_monitoring, daemon=True).start()
    
    def run_monitoring(self):
        """백그라운드에서 모니터링을 실행합니다."""
        while self.monitoring_active:
            try:
                # 현재 설정된 URL로 크롤링 실행
                url = self.url_var.get().strip()
                if url:
                    self.check_for_alerts(url)
                time.sleep(300)  # 5분마다 체크
            except Exception as e:
                print(f"모니터링 오류: {e}")
                time.sleep(60)  # 오류 시 1분 후 재시도
    
    def check_for_alerts(self, url):
        """알림 조건을 체크합니다."""
        try:
            # 간단한 크롤링으로 현재 데이터 확인
            response = requests.get(url, timeout=10)
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # 가격 정보 추출 (간단한 패턴)
            price_elements = soup.find_all(text=re.compile(r'[\d,]+원|[\$][\d,]+|\$[\d,.]+'))
            
            # 키워드 체크
            keywords = list(self.keyword_listbox.get(0, tk.END))
            page_text = soup.get_text().lower()
            
            alert_messages = []
            
            # 키워드 알림 체크
            for keyword in keywords:
                if keyword.lower() in page_text:
                    alert_messages.append(f"키워드 '{keyword}'가 발견되었습니다!")
            
            # 가격 변동 체크 (기본적인 구현)
            if price_elements and self.historical_data:
                current_prices = [self.extract_number(price.strip()) for price in price_elements[:3]]
                
                # 이전 데이터와 비교
                if len(self.historical_data) > 0:
                    prev_data = self.historical_data[-1]
                    threshold = float(self.price_threshold.get() or 10)
                    
                    for i, current_price in enumerate(current_prices):
                        if current_price and i < len(prev_data.get('prices', [])):
                            prev_price = prev_data['prices'][i]
                            if prev_price and current_price:
                                change_percent = abs((current_price - prev_price) / prev_price * 100)
                                if change_percent >= threshold:
                                    alert_messages.append(f"가격 변동 감지: {change_percent:.1f}% 변화")
            
            # 현재 데이터 저장
            current_data = {
                'timestamp': datetime.now(),
                'url': url,
                'prices': [self.extract_number(price.strip()) for price in price_elements[:3]],
                'keywords_found': [kw for kw in keywords if kw.lower() in page_text]
            }
            self.historical_data.append(current_data)
            
            # 데이터 제한 (최근 100개만 유지)
            if len(self.historical_data) > 100:
                self.historical_data = self.historical_data[-100:]
            
            # 알림 발송
            if alert_messages:
                self.send_alerts(alert_messages, url)
                
        except Exception as e:
            print(f"알림 체크 오류: {e}")
    
    def extract_number(self, text):
        """텍스트에서 숫자를 추출합니다."""
        try:
            # 숫자와 콤마만 추출
            numbers = re.findall(r'[\d,]+', text.replace(',', ''))
            if numbers:
                return float(numbers[0])
        except:
            pass
        return None
    
    def send_alerts(self, messages, url):
        """알림을 발송합니다."""
        alert_text = "\n".join(messages)
        alert_title = "웹 크롤링 알림"
        
        # 데스크탑 알림
        if self.notification_enabled.get() and NOTIFICATION_AVAILABLE:
            try:
                notification.notify(
                    title=alert_title,
                    message=alert_text,
                    timeout=10
                )
            except Exception as e:
                print(f"데스크탑 알림 오류: {e}")
        
        # 이메일 알림
        if self.email_enabled.get():
            self.send_email_alert(alert_title, alert_text, url)
    
    def send_email_alert(self, subject, message, url):
        """이메일 알림을 발송합니다."""
        try:
            sender_email = self.sender_email.get()
            sender_password = self.sender_password.get()
            receiver_email = self.receiver_email.get()
            
            if not all([sender_email, sender_password, receiver_email]):
                print("이메일 설정이 완료되지 않았습니다.")
                return
            
            # 이메일 메시지 구성
            msg = MimeMultipart()
            msg['From'] = sender_email
            msg['To'] = receiver_email
            msg['Subject'] = subject
            
            body = f"""
웹 크롤링 알림

URL: {url}
시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

알림 내용:
{message}

자동 발송된 메시지입니다.
            """
            
            msg.attach(MimeText(body, 'plain', 'utf-8'))
            
            # Gmail SMTP 서버 설정
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(sender_email, sender_password)
            
            server.send_message(msg)
            server.quit()
            
            print("이메일 알림 발송 완료")
            
        except Exception as e:
            print(f"이메일 발송 오류: {e}")
    
    def test_notification(self):
        """알림 테스트를 실행합니다."""
        test_messages = ["테스트 알림입니다.", "알림 설정이 정상적으로 작동합니다."]
        self.send_alerts(test_messages, "테스트 URL")
        messagebox.showinfo("테스트 완료", "알림 테스트가 완료되었습니다.")
    
    # 데이터 분석 관련 메서드들
    def generate_chart(self):
        """차트를 생성합니다."""
        if not CHART_AVAILABLE:
            messagebox.showerror("오류", "차트 기능을 사용하려면 matplotlib 라이브러리를 설치해주세요.")
            return
        
        if not self.historical_data and not self.crawled_data:
            messagebox.showwarning("경고", "분석할 데이터가 없습니다. 먼저 크롤링을 실행해주세요.")
            return
        
        chart_type = self.chart_type.get()
        
        try:
            # 기존 차트 제거
            for widget in self.chart_frame.winfo_children():
                widget.destroy()
            
            # matplotlib 한글 폰트 설정
            plt.rcParams['font.family'] = ['AppleGothic'] if os.name == 'posix' else ['Malgun Gothic']
            plt.rcParams['axes.unicode_minus'] = False
            
            fig, ax = plt.subplots(figsize=(8, 5))
            
            if chart_type == "가격 추이":
                self.create_price_trend_chart(ax)
            elif chart_type == "키워드 빈도":
                self.create_keyword_frequency_chart(ax)
            elif chart_type == "사이트별 통계":
                self.create_site_statistics_chart(ax)
            elif chart_type == "시간대별 분포":
                self.create_time_distribution_chart(ax)
            
            # 차트를 tkinter에 표시
            canvas = FigureCanvasTkAgg(fig, self.chart_frame)
            canvas.draw()
            canvas.get_tk_widget().grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
            
        except Exception as e:
            messagebox.showerror("차트 생성 오류", f"차트 생성 중 오류가 발생했습니다:\n{str(e)}")
    
    def create_price_trend_chart(self, ax):
        """가격 추이 차트를 생성합니다."""
        if not self.historical_data:
            ax.text(0.5, 0.5, '가격 데이터가 없습니다', ha='center', va='center', transform=ax.transAxes)
            return
        
        timestamps = [data['timestamp'] for data in self.historical_data]
        prices = []
        
        for data in self.historical_data:
            if data['prices'] and data['prices'][0]:
                prices.append(data['prices'][0])
            else:
                prices.append(0)
        
        ax.plot(timestamps, prices, marker='o', linewidth=2, markersize=4)
        ax.set_title('가격 추이')
        ax.set_xlabel('시간')
        ax.set_ylabel('가격')
        ax.grid(True, alpha=0.3)
        
        # x축 날짜 포맷 설정
        import matplotlib.dates as mdates
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%d %H:%M'))
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=45)
    
    def create_keyword_frequency_chart(self, ax):
        """키워드 빈도 차트를 생성합니다."""
        keyword_counts = {}
        
        # 크롤링 데이터에서 키워드 빈도 계산
        for data in self.crawled_data:
            title = data.get('title', '').lower()
            description = data.get('description', '').lower()
            text = title + ' ' + description
            
            # 간단한 단어 분리 (한글/영문)
            words = re.findall(r'[가-힣]{2,}|[a-zA-Z]{3,}', text)
            for word in words:
                keyword_counts[word] = keyword_counts.get(word, 0) + 1
        
        if not keyword_counts:
            ax.text(0.5, 0.5, '키워드 데이터가 없습니다', ha='center', va='center', transform=ax.transAxes)
            return
        
        # 상위 10개 키워드
        top_keywords = sorted(keyword_counts.items(), key=lambda x: x[1], reverse=True)[:10]
        keywords, counts = zip(*top_keywords)
        
        ax.bar(keywords, counts)
        ax.set_title('키워드 빈도')
        ax.set_xlabel('키워드')
        ax.set_ylabel('빈도')
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=45)
    
    def create_site_statistics_chart(self, ax):
        """사이트별 통계 차트를 생성합니다."""
        site_counts = {}
        
        for data in self.crawled_data:
            url = data.get('url', '')
            domain = urlparse(url).netloc
            site_counts[domain] = site_counts.get(domain, 0) + 1
        
        if not site_counts:
            ax.text(0.5, 0.5, '사이트 데이터가 없습니다', ha='center', va='center', transform=ax.transAxes)
            return
        
        sites, counts = zip(*site_counts.items())
        ax.pie(counts, labels=sites, autopct='%1.1f%%')
        ax.set_title('사이트별 데이터 분포')
    
    def create_time_distribution_chart(self, ax):
        """시간대별 분포 차트를 생성합니다."""
        if not self.historical_data:
            ax.text(0.5, 0.5, '시간 데이터가 없습니다', ha='center', va='center', transform=ax.transAxes)
            return
        
        hours = [data['timestamp'].hour for data in self.historical_data]
        hour_counts = {}
        for hour in hours:
            hour_counts[hour] = hour_counts.get(hour, 0) + 1
        
        hours_list = list(range(24))
        counts_list = [hour_counts.get(hour, 0) for hour in hours_list]
        
        ax.bar(hours_list, counts_list)
        ax.set_title('시간대별 활동 분포')
        ax.set_xlabel('시간')
        ax.set_ylabel('활동 횟수')
        ax.set_xticks(range(0, 24, 2))
    
    def generate_report(self):
        """분석 리포트를 생성합니다."""
        if not self.crawled_data and not self.historical_data:
            messagebox.showwarning("경고", "분석할 데이터가 없습니다.")
            return
        
        # 통계 계산
        total_items = len(self.crawled_data)
        total_monitoring = len(self.historical_data)
        
        # 사이트별 통계
        site_stats = {}
        for data in self.crawled_data:
            site_type = data.get('type', '기타')
            site_stats[site_type] = site_stats.get(site_type, 0) + 1
        
        # 리포트 텍스트 생성
        report = f"""크롤링 데이터 분석 리포트
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

생성 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

📊 전체 통계
• 총 크롤링 항목: {total_items}개
• 모니터링 기록: {total_monitoring}개

📈 사이트별 분포
"""
        
        for site_type, count in site_stats.items():
            percentage = (count / total_items * 100) if total_items > 0 else 0
            report += f"• {site_type}: {count}개 ({percentage:.1f}%)\n"
        
        if self.historical_data:
            report += f"""
⏰ 모니터링 정보
• 첫 기록: {self.historical_data[0]['timestamp'].strftime('%Y-%m-%d %H:%M')}
• 마지막 기록: {self.historical_data[-1]['timestamp'].strftime('%Y-%m-%d %H:%M')}
• 총 모니터링 기간: {(self.historical_data[-1]['timestamp'] - self.historical_data[0]['timestamp']).days}일
"""
        
        # 키워드 분석
        if self.keyword_listbox.size() > 0:
            keywords = list(self.keyword_listbox.get(0, tk.END))
            report += f"""
🔍 모니터링 키워드
• 설정된 키워드: {', '.join(keywords)}
"""
        
        report += f"""
📋 분석 권장사항
• 데이터가 충분히 쌓이면 더 정확한 트렌드 분석이 가능합니다
• 정기적인 모니터링으로 변화 패턴을 파악하세요
• 알림 설정을 통해 중요한 변화를 놓치지 마세요

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""
        
        # 통계 텍스트 영역에 표시
        self.stats_text.delete(1.0, tk.END)
        self.stats_text.insert(1.0, report)
        
        messagebox.showinfo("리포트 생성 완료", "분석 리포트가 생성되었습니다.")
    
    # 스케줄링 관련 메서드들
    def add_schedule(self):
        """새로운 스케줄을 추가합니다."""
        schedule_type = self.schedule_type.get()
        hour = int(self.schedule_hour.get())
        minute = int(self.schedule_minute.get())
        
        try:
            if schedule_type == "일회성":
                # 오늘 또는 내일 지정된 시간에 실행
                now = datetime.now()
                scheduled_time = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
                if scheduled_time <= now:
                    scheduled_time += timedelta(days=1)
                
                job = self.scheduler.add_job(
                    func=self.scheduled_crawl,
                    trigger='date',
                    run_date=scheduled_time,
                    id=f"onetime_{scheduled_time.strftime('%Y%m%d_%H%M%S')}"
                )
                
            elif schedule_type == "매일":
                job = self.scheduler.add_job(
                    func=self.scheduled_crawl,
                    trigger=CronTrigger(hour=hour, minute=minute),
                    id=f"daily_{hour:02d}{minute:02d}"
                )
                
            elif schedule_type == "매주":
                weekday_map = {"월요일": 0, "화요일": 1, "수요일": 2, "목요일": 3, 
                             "금요일": 4, "토요일": 5, "일요일": 6}
                weekday = weekday_map[self.schedule_weekday.get()]
                
                job = self.scheduler.add_job(
                    func=self.scheduled_crawl,
                    trigger=CronTrigger(day_of_week=weekday, hour=hour, minute=minute),
                    id=f"weekly_{weekday}_{hour:02d}{minute:02d}"
                )
                
            elif schedule_type == "매월":
                day = int(self.schedule_day.get())
                job = self.scheduler.add_job(
                    func=self.scheduled_crawl,
                    trigger=CronTrigger(day=day, hour=hour, minute=minute),
                    id=f"monthly_{day}_{hour:02d}{minute:02d}"
                )
            
            self.scheduled_jobs.append({
                'job': job,
                'type': schedule_type,
                'time': f"{hour:02d}:{minute:02d}",
                'email_results': self.email_results.get()
            })
            
            self.update_schedule_list()
            messagebox.showinfo("스케줄 추가", f"{schedule_type} 스케줄이 추가되었습니다.")
            
        except Exception as e:
            messagebox.showerror("스케줄 추가 오류", f"스케줄 추가 중 오류가 발생했습니다:\n{str(e)}")
    
    def remove_schedule(self):
        """선택된 스케줄을 삭제합니다."""
        try:
            selection = self.schedule_tree.selection()
            if not selection:
                messagebox.showwarning("선택 필요", "삭제할 스케줄을 선택해주세요.")
                return
            
            item = self.schedule_tree.item(selection[0])
            job_id = item['values'][0]
            
            # 스케줄러에서 작업 제거
            self.scheduler.remove_job(job_id)
            
            # 목록에서 제거
            self.scheduled_jobs = [job for job in self.scheduled_jobs if job['job'].id != job_id]
            
            self.update_schedule_list()
            messagebox.showinfo("스케줄 삭제", "선택된 스케줄이 삭제되었습니다.")
            
        except Exception as e:
            messagebox.showerror("스케줄 삭제 오류", f"스케줄 삭제 중 오류가 발생했습니다:\n{str(e)}")
    
    def update_schedule_list(self):
        """스케줄 목록을 업데이트합니다."""
        # 기존 항목 삭제
        for item in self.schedule_tree.get_children():
            self.schedule_tree.delete(item)
        
        # 현재 스케줄 목록 추가
        for job_info in self.scheduled_jobs:
            job = job_info['job']
            next_run = job.next_run_time.strftime('%Y-%m-%d %H:%M:%S') if job.next_run_time else "N/A"
            
            self.schedule_tree.insert('', 'end', values=(
                job.id,
                job_info['type'],
                job_info['time'],
                next_run,
                "활성"
            ))
    
    def scheduled_crawl(self):
        """스케줄된 크롤링을 실행합니다."""
        try:
            # 스케줄된 크롤링 실행
            self.start_crawling(scheduled=True)
            
            # 크롤링 완료까지 대기 (최대 30분)
            wait_time = 0
            while self.is_crawling and wait_time < 1800:  # 30분
                time.sleep(10)
                wait_time += 10
            
            # 결과 이메일 발송
            if self.email_results.get() and self.email_enabled.get():
                self.send_crawling_results_email()
                
        except Exception as e:
            print(f"스케줄된 크롤링 오류: {e}")
    
    def send_crawling_results_email(self):
        """크롤링 결과를 이메일로 전송합니다."""
        try:
            if not self.crawled_data:
                return
            
            # 결과 요약 생성
            total_items = len(self.crawled_data)
            sites = set(urlparse(item.get('url', '')).netloc for item in self.crawled_data)
            
            # 엑셀 파일 생성
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_filename = f"scheduled_crawling_{timestamp}.xlsx"
            
            # 임시로 엑셀 파일 저장
            processed_data = []
            for item in self.crawled_data:
                processed_item = {}
                for key, value in item.items():
                    if isinstance(value, str):
                        value = unicodedata.normalize('NFC', value)
                        value = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', value)
                        value = re.sub(r'\s+', ' ', value).strip()
                        if len(value) > 32000:
                            value = value[:32000] + "..."
                    processed_item[key] = value
                processed_data.append(processed_item)
            
            df = pd.DataFrame(processed_data)
            df.columns = ['타입', '제목/텍스트', 'URL', '설명']
            
            with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='CrawlingResults', index=False)
            
            # 이메일 내용 구성
            subject = f"웹 크롤링 결과 - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            
            body = f"""
자동 웹 크롤링 결과

실행 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
크롤링 URL: {self.current_task['url'] if self.current_task else 'N/A'}

📊 결과 요약:
• 총 수집 항목: {total_items}개
• 크롤링 사이트: {len(sites)}개
• 파일명: {excel_filename}

상세 결과는 첨부된 엑셀 파일을 확인해주세요.

자동 발송된 메시지입니다.
            """
            
            # 파일 첨부 이메일 발송
            self.send_email_with_attachment(subject, body, excel_filename)
            
            # 임시 파일 삭제
            if os.path.exists(excel_filename):
                os.remove(excel_filename)
                
        except Exception as e:
            print(f"결과 이메일 발송 오류: {e}")
    
    def send_email_with_attachment(self, subject, body, filename):
        """첨부파일이 있는 이메일을 발송합니다."""
        try:
            from email.mime.base import MimeBase
            from email import encoders
            
            sender_email = self.sender_email.get()
            sender_password = self.sender_password.get()
            receiver_email = self.receiver_email.get()
            
            if not all([sender_email, sender_password, receiver_email]):
                print("이메일 설정이 완료되지 않았습니다.")
                return
            
            msg = MimeMultipart()
            msg['From'] = sender_email
            msg['To'] = receiver_email
            msg['Subject'] = subject
            
            # 본문 추가
            msg.attach(MimeText(body, 'plain', 'utf-8'))
            
            # 파일 첨부
            if os.path.exists(filename):
                with open(filename, "rb") as attachment:
                    part = MimeBase('application', 'octet-stream')
                    part.set_payload(attachment.read())
                
                encoders.encode_base64(part)
                part.add_header(
                    'Content-Disposition',
                    f'attachment; filename= {os.path.basename(filename)}'
                )
                msg.attach(part)
            
            # 이메일 발송
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(sender_email, sender_password)
            server.send_message(msg)
            server.quit()
            
            print("첨부파일 이메일 발송 완료")
            
        except Exception as e:
            print(f"첨부파일 이메일 발송 오류: {e}")
    
    # 체크포인트 및 복구 관련 메서드들
    def save_checkpoint(self):
        """현재 진행 상황을 체크포인트로 저장합니다."""
        try:
            checkpoint_data = {
                'task_progress': self.task_progress,
                'current_task': self.current_task,
                'crawled_data': self.crawled_data,
                'timestamp': datetime.now(),
                'current_page': self.current_page,
                'retry_count': self.retry_count
            }
            
            with open(self.checkpoint_file, 'wb') as f:
                pickle.dump(checkpoint_data, f)
            
            self.update_checkpoint_status()
            
        except Exception as e:
            print(f"체크포인트 저장 오류: {e}")
    
    def load_checkpoint(self):
        """저장된 체크포인트를 로드합니다."""
        try:
            if not os.path.exists(self.checkpoint_file):
                return None
            
            with open(self.checkpoint_file, 'rb') as f:
                checkpoint_data = pickle.load(f)
            
            return checkpoint_data
            
        except Exception as e:
            print(f"체크포인트 로드 오류: {e}")
            return None
    
    def restore_last_task(self):
        """마지막 작업을 복구합니다."""
        checkpoint_data = self.load_checkpoint()
        
        if not checkpoint_data:
            messagebox.showwarning("복구 불가", "복구할 체크포인트가 없습니다.")
            return
        
        try:
            # 체크포인트 데이터 복원
            self.task_progress = checkpoint_data['task_progress']
            self.current_task = checkpoint_data['current_task']
            self.crawled_data = checkpoint_data['crawled_data']
            self.current_page = checkpoint_data.get('current_page', 1)
            self.retry_count = checkpoint_data.get('retry_count', 0)
            
            # UI 설정 복원
            if self.current_task:
                self.url_var.set(self.current_task['url'])
                self.browser_mode.set(self.current_task['browser_mode'])
                self.max_pages.set(str(self.current_task['max_pages']))
                self.crawl_delay.set(str(self.current_task['crawl_delay']))
                self.site_type.set(self.current_task['site_type'])
            
            # 진행 상황 표시
            completed = self.task_progress['completed_pages']
            total = self.task_progress['total_pages']
            
            result = messagebox.askyesno(
                "작업 복구", 
                f"마지막 작업을 발견했습니다.\n\n"
                f"URL: {self.current_task['url']}\n"
                f"진행률: {completed}/{total} 페이지\n"
                f"시작 시간: {self.current_task['started_at'].strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                f"이 작업을 이어서 실행하시겠습니까?"
            )
            
            if result:
                # 크롤링 재시작
                self.resume_crawling()
            
        except Exception as e:
            messagebox.showerror("복구 오류", f"작업 복구 중 오류가 발생했습니다:\n{str(e)}")
    
    def resume_crawling(self):
        """중단된 크롤링을 재시작합니다."""
        try:
            self.is_crawling = True
            self.status_var.set(f"복구된 작업 재시작 중... ({self.task_progress['completed_pages']}/{self.task_progress['total_pages']})")
            
            # UI 상태 변경
            self.crawl_button.config(state='disabled')
            self.stop_button.config(state='normal')
            self.progress.start()
            
            # 실패한 페이지부터 재시작
            url = self.current_task['url']
            
            if self.current_task['browser_mode'] == "selenium":
                thread = threading.Thread(target=self.crawl_with_selenium_checkpoint, args=(url, False, True))
            elif self.current_task['browser_mode'] == "playwright":
                thread = threading.Thread(target=self.crawl_with_playwright_checkpoint, args=(url, False, True))
            else:
                thread = threading.Thread(target=self.crawl_website_checkpoint, args=(url, False, True))
            
            thread.daemon = True
            thread.start()
            
        except Exception as e:
            messagebox.showerror("재시작 오류", f"크롤링 재시작 중 오류가 발생했습니다:\n{str(e)}")
    
    def clear_checkpoint(self, show_message=True):
        """체크포인트를 삭제합니다."""
        try:
            if os.path.exists(self.checkpoint_file):
                os.remove(self.checkpoint_file)
            
            self.task_progress = {
                'total_pages': 0,
                'completed_pages': 0,
                'failed_pages': [],
                'current_url': '',
                'settings': {}
            }
            self.current_task = None
            
            self.update_checkpoint_status()
            
            # 메시지 표시 여부를 선택적으로 처리
            if show_message:
                messagebox.showinfo("체크포인트 삭제", "체크포인트가 삭제되었습니다.")
            
        except Exception as e:
            if show_message:
                messagebox.showerror("삭제 오류", f"체크포인트 삭제 중 오류가 발생했습니다:\n{str(e)}")
            else:
                print(f"체크포인트 삭제 오류: {str(e)}")
    
    def update_checkpoint_status(self):
        """체크포인트 상태를 업데이트합니다."""
        try:
            if os.path.exists(self.checkpoint_file):
                checkpoint_data = self.load_checkpoint()
                if checkpoint_data:
                    timestamp = checkpoint_data['timestamp'].strftime('%Y-%m-%d %H:%M:%S')
                    completed = checkpoint_data['task_progress']['completed_pages']
                    total = checkpoint_data['task_progress']['total_pages']
                    self.checkpoint_status.set(f"체크포인트 있음: {timestamp} ({completed}/{total})")
                else:
                    self.checkpoint_status.set("체크포인트 파일 손상됨")
            else:
                self.checkpoint_status.set("체크포인트 없음")
        except:
            self.checkpoint_status.set("체크포인트 상태 확인 불가")
    
    # 체크포인트를 포함한 크롤링 메서드들
    def crawl_website_checkpoint(self, url, scheduled=False, resume=False):
        """체크포인트 기능이 포함된 기본 크롤링"""
        try:
            start_page = self.current_page if resume else 1
            max_pages = int(self.max_pages.get()) if self.max_pages.get().isdigit() else 1
            
            for page in range(start_page, max_pages + 1):
                if not self.is_crawling:
                    break
                
                self.current_page = page
                self.task_progress['completed_pages'] = page - 1
                
                # 체크포인트 저장
                if self.auto_save.get() and page % 5 == 0:  # 5페이지마다 저장
                    self.save_checkpoint()
                
                # 페이지 크롤링 시도
                success = False
                for retry in range(self.max_retries):
                    try:
                        page_url = self.generate_page_url(url, page)
                        response = self.crawl_with_retry(page_url)
                        
                        if response:
                            soup = BeautifulSoup(response.content, 'html.parser')
                            self.root.after(0, self.update_results, page_url, response, soup)
                            success = True
                            break
                            
                    except Exception as e:
                        self.root.after(0, lambda r=retry+1, err=str(e): self.status_var.set(f"페이지 {page} 오류, 재시도 {r}/{self.max_retries}: {err[:30]}"))
                        time.sleep(2)
                
                if not success:
                    self.task_progress['failed_pages'].append(page)
                
                # 크롤링 간격
                if page < max_pages and self.is_crawling:
                    time.sleep(float(self.crawl_delay.get()) if self.crawl_delay.get().replace('.', '').isdigit() else 1)
            
            # 작업 완료 처리
            self.root.after(0, self.finalize_crawling_checkpoint, scheduled)
            
        except Exception as e:
            self.root.after(0, self.show_error, f"체크포인트 크롤링 오류: {str(e)}")
    
    def crawl_with_selenium_checkpoint(self, url, scheduled=False, resume=False):
        """체크포인트 기능이 포함된 Selenium 크롤링"""
        if not self.setup_selenium_driver():
            return
        
        try:
            start_page = self.current_page if resume else 1
            max_pages = int(self.max_pages.get()) if self.max_pages.get().isdigit() else 1
            
            for page in range(start_page, max_pages + 1):
                if not self.is_crawling:
                    break
                
                self.current_page = page
                self.task_progress['completed_pages'] = page - 1
                
                # 체크포인트 저장
                if self.auto_save.get() and page % 3 == 0:  # 3페이지마다 저장
                    self.save_checkpoint()
                
                success = False
                for retry in range(self.max_retries):
                    try:
                        page_url = self.generate_page_url(url, page)
                        self.driver.get(page_url)
                        WebDriverWait(self.driver, 10).until(
                            EC.presence_of_element_located((By.TAG_NAME, "body"))
                        )
                        
                        # 사이트별 특화 크롤링
                        if self.site_type.get() == "네이버 쇼핑":
                            self.crawl_naver_shopping()
                        elif self.site_type.get() == "인스타그램":
                            self.crawl_instagram()
                        elif self.site_type.get() == "부동산":
                            self.crawl_real_estate()
                        else:
                            self.crawl_general_selenium()
                        
                        success = True
                        break
                        
                    except Exception as e:
                        self.root.after(0, lambda r=retry+1, err=str(e): self.status_var.set(f"Selenium 페이지 {page} 재시도 {r}/{self.max_retries}: {err[:30]}"))
                        time.sleep(2)
                
                if not success:
                    self.task_progress['failed_pages'].append(page)
                
                if page < max_pages and self.is_crawling:
                    time.sleep(float(self.crawl_delay.get()) if self.crawl_delay.get().replace('.', '').isdigit() else 1)
            
            self.root.after(0, self.finalize_crawling_checkpoint, scheduled)
            
        except Exception as e:
            self.root.after(0, self.show_error, f"Selenium 체크포인트 크롤링 오류: {str(e)}")
        finally:
            if self.driver:
                try:
                    self.driver.quit()
                except:
                    pass
                self.driver = None
    
    def crawl_with_playwright_checkpoint(self, url, scheduled=False, resume=False):
        """체크포인트 기능이 포함된 Playwright 크롤링"""
        if not self.setup_playwright():
            return
        
        try:
            start_page = self.current_page if resume else 1
            max_pages = int(self.max_pages.get()) if self.max_pages.get().isdigit() else 1
            
            for page in range(start_page, max_pages + 1):
                if not self.is_crawling:
                    break
                
                self.current_page = page
                self.task_progress['completed_pages'] = page - 1
                
                # 체크포인트 저장
                if self.auto_save.get() and page % 3 == 0:  # 3페이지마다 저장
                    self.save_checkpoint()
                
                success = False
                for retry in range(self.max_retries):
                    try:
                        page_url = self.generate_page_url(url, page)
                        self.page.goto(page_url, wait_until='domcontentloaded', timeout=30000)
                        self.page.wait_for_load_state('networkidle', timeout=10000)
                        self.page.wait_for_timeout(2000)
                        
                        domain = urlparse(page_url).netloc.lower()
                        if 'shopping.naver.com' in domain:
                            self.crawl_naver_shopping_playwright()
                        elif 'instagram.com' in domain:
                            self.crawl_instagram_playwright()
                        elif any(keyword in domain for keyword in ['zigbang', 'dabang', '부동산']):
                            self.crawl_real_estate_playwright()
                        else:
                            self.crawl_general_playwright()
                        
                        success = True
                        break
                        
                    except Exception as e:
                        self.root.after(0, lambda r=retry+1, err=str(e): self.status_var.set(f"Playwright 페이지 {page} 재시도 {r}/{self.max_retries}: {err[:30]}"))
                        time.sleep(2)
                
                if not success:
                    self.task_progress['failed_pages'].append(page)
                
                if page < max_pages and self.is_crawling:
                    time.sleep(float(self.crawl_delay.get()) if self.crawl_delay.get().replace('.', '').isdigit() else 1)
            
            self.root.after(0, self.finalize_crawling_checkpoint, scheduled)
            
        except Exception as e:
            self.root.after(0, self.show_error, f"Playwright 체크포인트 크롤링 오류: {str(e)}")
        finally:
            # Playwright 정리
            if self.page:
                try:
                    self.page.close()
                except:
                    pass
                self.page = None
            
            if self.browser:
                try:
                    self.browser.close()
                except:
                    pass
                self.browser = None
            
            if self.playwright:
                try:
                    self.playwright.stop()
                except:
                    pass
                self.playwright = None
    
    def finalize_crawling_checkpoint(self, scheduled=False):
        """체크포인트 기능이 포함된 크롤링 완료 처리"""
        self.progress.stop()
        if not scheduled:
            self.crawl_button.config(state='normal')
            self.stop_button.config(state='disabled')
            self.export_button.config(state='normal')
        
        self.is_crawling = False
        
        total_items = len(self.crawled_data)
        failed_pages = len(self.task_progress['failed_pages'])
        
        status_msg = f"크롤링 완료 - 총 {total_items}개 데이터 수집"
        if failed_pages > 0:
            status_msg += f" (실패: {failed_pages}페이지)"
        
        self.status_var.set(status_msg)
        
        # 체크포인트 정리 (작업 완료시) - 메시지 없이 자동 삭제
        if self.auto_save.get():
            self.clear_checkpoint(show_message=False)

def main():
    root = tk.Tk()
    app = WebCrawlerApp(root)
    root.mainloop()

if __name__ == "__main__":
    main() 